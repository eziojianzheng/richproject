#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
PostgreSQL 数据库模块

用途: 把 excelDataSource 里的涨停复盘 Excel 数据提交(入库)到 PostgreSQL，
      供热门股追踪等后续分析使用。

连接配置优先级: config.yml 的 database 段 > 默认值(localhost:15432 postgres/postgres)

表结构:
  zt_daily  (每日涨跌/成交, 一天一行, 兼作"已入库"标记)
  zt_stocks (每日板块个股明细, 一天多行)
"""

import os
import re
import glob

try:
    import yaml
except Exception:
    yaml = None

import openpyxl

# psycopg2 延迟导入，未安装时给出清晰提示
try:
    import psycopg2
    from psycopg2.extras import execute_values
    _PSYCOPG_OK = True
except Exception:
    psycopg2 = None
    execute_values = None
    _PSYCOPG_OK = False


EXCEL_DIR = 'excelDataSource'

# 04提取 sheet 末尾的校验说明行(非个股), 入库时跳过
_META_FIRST_COL = {'【校验结果】', '问题1', '问题2', '问题3', '问题4',
                   '问题5', '问题6', '问题7', '建议', '说明', '【未找到03图】', '【未找到04图】'}

_DEFAULT_DB = {
    'host': 'localhost',
    'port': 15432,
    'user': 'postgres',
    'password': 'postgres',
    'dbname': 'postgres',
}


class DBError(Exception):
    """数据库相关错误(连接失败/驱动缺失等)"""
    pass


def _load_db_config():
    cfg = dict(_DEFAULT_DB)
    if yaml and os.path.exists('config.yml'):
        try:
            with open('config.yml', 'r', encoding='utf-8') as f:
                data = yaml.safe_load(f) or {}
            db = data.get('database') or {}
            for k in cfg:
                if db.get(k) is not None:
                    cfg[k] = db[k]
        except Exception:
            pass
    return cfg


def get_conn(timeout=5):
    """获取数据库连接；驱动缺失或连不上时抛 DBError"""
    if not _PSYCOPG_OK:
        raise DBError("未安装 psycopg2，请执行: pip install psycopg2-binary")
    cfg = _load_db_config()
    try:
        return psycopg2.connect(
            host=cfg['host'], port=int(cfg['port']),
            user=cfg['user'], password=cfg['password'],
            dbname=cfg['dbname'], connect_timeout=timeout,
        )
    except Exception as e:
        raise DBError(f"数据库连接失败({cfg['host']}:{cfg['port']}): {e}")


def ping():
    """探测数据库是否可用，返回 (ok: bool, msg: str)"""
    try:
        conn = get_conn()
        conn.close()
        return True, "ok"
    except DBError as e:
        return False, str(e)


def init_db():
    """建表(幂等)"""
    ddl = """
    CREATE TABLE IF NOT EXISTS zt_daily (
        trade_date   DATE PRIMARY KEY,
        up_count     INTEGER,
        down_count   INTEGER,
        total_amount NUMERIC,
        status       TEXT,
        updated_at   TIMESTAMP DEFAULT now()
    );
    CREATE TABLE IF NOT EXISTS zt_stocks (
        id         SERIAL PRIMARY KEY,
        trade_date DATE NOT NULL,
        block      TEXT,
        code       TEXT,
        name       TEXT,
        last_time  TEXT,
        lianban    TEXT,
        reason     TEXT
    );
    CREATE INDEX IF NOT EXISTS idx_zt_stocks_date ON zt_stocks(trade_date);
    """
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(ddl)
        conn.commit()
    finally:
        conn.close()


# ============== Excel 解析 ==============

def _date_from_filename(fp):
    m = re.search(r'(\d{8})', os.path.basename(fp))
    return m.group(1) if m else None


def find_excel(date, excel_dir=EXCEL_DIR):
    """返回某日期的 Excel 路径与状态: (path|None, status)  status: verified/manualcheck/None"""
    verified = os.path.join(excel_dir, f'{date}_涨停复盘_verified.xlsx')
    manual = os.path.join(excel_dir, f'{date}_涨停复盘_manualcheck.xlsx')
    if os.path.exists(verified):
        return verified, 'verified'
    if os.path.exists(manual):
        return manual, 'manualcheck'
    return None, None


def _num(v):
    """把 '4329' / '36827.97' 之类转成数字，失败返回 None"""
    if v is None:
        return None
    s = str(v).strip().replace(',', '')
    if s == '':
        return None
    try:
        return float(s) if '.' in s else int(s)
    except ValueError:
        m = re.search(r'-?\d+(\.\d+)?', s)
        return float(m.group(0)) if m else None


def parse_excel(fp):
    """
    解析单个 Excel，返回 (daily: dict, stocks: list[dict])
      daily: {up_count, down_count, total_amount}
      stocks: [{block, code, name, last_time, lianban, reason}, ...]
    """
    wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    daily = {'up_count': None, 'down_count': None, 'total_amount': None}

    if '03提取' in wb.sheetnames:
        ws = wb['03提取']
        row2 = None
        for r in ws.iter_rows(min_row=2, max_row=2, values_only=True):
            row2 = r
            break
        if row2:
            daily['up_count'] = _num(row2[0]) if len(row2) > 0 else None
            daily['down_count'] = _num(row2[1]) if len(row2) > 1 else None
            daily['total_amount'] = _num(row2[2]) if len(row2) > 2 else None

    stocks = []
    if '04提取' in wb.sheetnames:
        ws = wb['04提取']
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or not r[0]:
                continue
            block = str(r[0]).strip()
            if block in _META_FIRST_COL:
                continue
            code = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ''
            if not re.match(r'^\d{6}$', code):
                continue
            stocks.append({
                'block': block,
                'code': code,
                'name': str(r[2]).strip() if len(r) > 2 and r[2] is not None else '',
                'last_time': str(r[3]).strip() if len(r) > 3 and r[3] is not None else '',
                'lianban': str(r[4]).strip() if len(r) > 4 and r[4] is not None else '',
                'reason': str(r[5]).strip() if len(r) > 5 and r[5] is not None else '',
            })
    wb.close()
    return daily, stocks


def _fmt_date(date8):
    """20260701 -> 2026-07-01"""
    return f"{date8[:4]}-{date8[4:6]}-{date8[6:8]}"


# ============== 入库 ==============

def submit_date(date, excel_dir=EXCEL_DIR):
    """
    把某日期的 Excel 数据入库(覆盖式: 先删该日期旧数据再插入)。
    返回 dict: {date, status, stocks, up_count, down_count, total_amount}
    找不到 Excel 抛 FileNotFoundError；数据库错误抛 DBError。
    """
    date = str(date).replace('-', '')
    fp, status = find_excel(date, excel_dir)
    if not fp:
        raise FileNotFoundError(f"未找到 {date} 的 Excel(verified/manualcheck)")

    daily, stocks = parse_excel(fp)
    d = _fmt_date(date)

    conn = get_conn()
    try:
        with conn.cursor() as cur:
            # 覆盖式: 先删旧
            cur.execute("DELETE FROM zt_stocks WHERE trade_date = %s", (d,))
            cur.execute("DELETE FROM zt_daily WHERE trade_date = %s", (d,))
            # 03
            cur.execute(
                "INSERT INTO zt_daily (trade_date, up_count, down_count, total_amount, status) "
                "VALUES (%s, %s, %s, %s, %s)",
                (d, daily['up_count'], daily['down_count'], daily['total_amount'], status),
            )
            # 04
            if stocks:
                execute_values(
                    cur,
                    "INSERT INTO zt_stocks (trade_date, block, code, name, last_time, lianban, reason) VALUES %s",
                    [(d, s['block'], s['code'], s['name'], s['last_time'], s['lianban'], s['reason'])
                     for s in stocks],
                )
        conn.commit()
    finally:
        conn.close()

    return {
        'date': date, 'status': status, 'stocks': len(stocks),
        'up_count': daily['up_count'], 'down_count': daily['down_count'],
        'total_amount': daily['total_amount'],
    }


def get_submitted_dates():
    """返回已入库的日期集合(YYYYMMDD)"""
    conn = get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT to_char(trade_date, 'YYYYMMDD') FROM zt_daily")
            return {r[0] for r in cur.fetchall()}
    finally:
        conn.close()
