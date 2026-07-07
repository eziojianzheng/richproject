#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
热门股追踪 - 数据统计核心模块

数据源: excelDataSource/<日期>_涨停复盘_*.xlsx 的「04提取」sheet
列: 概念板块 | 代码 | 名称 | 末次时间 | 连扳数 | 原因

规则:
  统计范围: 仅概念板块, 排除「市场连板股」
  板块入选(满足任一即加入显示列表, 首次满足后一直保留):
      - 板块内存在 连板>=2 的票
      - 板块内存在 "1/X板" 格式的票 (如 1/3天2板)
      - 板块内存在 9:25 涨停(一字板)的票
  个股入选(已入选板块下, 满足任一):
      - "1/X板" 格式
      - 9:25 涨停
      - 300/688 开头的 1 板
  入选个股每日跟踪: 在日期范围内逐日展示其连板进展/涨幅
  排序: 可切换 'stock_count'(板块内入选票数) / 'days'(板块出现天数)
"""

import sys
try:
    sys.stdout.reconfigure(encoding='utf-8')
except Exception:
    pass

import logging as _logging
_ht_logger = _logging.getLogger('ai_kanpan')

import os
import re
import glob
import json

import openpyxl


EXCEL_DIR = 'excelDataSource'
EXCLUDED_BLOCKS = {'市场连板股'}
# 04提取 sheet 末尾的校验说明行, 解析时跳过
_META_FIRST_COL = {'【校验结果】', '问题1', '问题2', '问题3', '问题4', '建议', '说明'}


# ============== Excel 解析 ==============

def _date_from_filename(fp):
    """从文件名提取 8 位日期, 如 20260601_涨停复盘_verified.xlsx -> 20260601"""
    m = re.search(r'(\d{8})', os.path.basename(fp))
    return m.group(1) if m else None


def list_excel_dates(excel_dir=EXCEL_DIR):
    """返回 {日期: 文件路径}, 同一日期取任意一个(verified优先)"""
    mapping = {}
    for fp in glob.glob(os.path.join(excel_dir, '*_涨停复盘*.xlsx')):
        if os.path.basename(fp).startswith('~$'):
            continue  # Excel 临时锁文件
        d = _date_from_filename(fp)
        if not d:
            continue
        # verified 优先于 manualcheck
        if d not in mapping or 'verified' in os.path.basename(fp):
            mapping[d] = fp
    return mapping


def parse_excel_04(fp):
    """
    解析单个 Excel 的 04提取 sheet

    返回: list[dict] 每条 {概念板块, 代码, 名称, 末次时间, 连扳数, 原因}
          已排除末尾校验说明行
    """
    wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
    if '04提取' not in wb.sheetnames:
        return []
    ws = wb['04提取']
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0]:
            continue
        block = str(r[0]).strip()
        if block in _META_FIRST_COL:
            continue
        code = str(r[1]).strip() if len(r) > 1 and r[1] is not None else ''
        # 必须是 6 位股票代码才算有效个股行
        if not re.match(r'^\d{6}$', code):
            continue
        rows.append({
            '概念板块': block,
            '代码': code,
            '名称': str(r[2]).strip() if len(r) > 2 and r[2] is not None else '',
            '末次时间': str(r[3]).strip() if len(r) > 3 and r[3] is not None else '',
            '连扳数': str(r[4]).strip() if len(r) > 4 and r[4] is not None else '',
            '原因': str(r[5]).strip() if len(r) > 5 and r[5] is not None else '',
        })
    wb.close()
    return rows


# ============== 数据库数据源 (PostgreSQL) ==============

import db as _dbmod


def list_db_dates():
    """返回数据库中已入库的日期集合(YYYYMMDD)"""
    try:
        return set(_dbmod.get_submitted_dates())
    except Exception as e:
        print(f'读取数据库日期失败: {e}')
        return set()


def read_day_stocks_db(date8):
    """
    从 zt_stocks 读取某日个股, 返回与 parse_excel_04 相同结构的列表
    [{概念板块, 代码, 名称, 末次时间, 连扳数, 原因}, ...]
    """
    d = f"{date8[:4]}-{date8[4:6]}-{date8[6:8]}"
    conn = _dbmod.get_conn()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT block, code, name, last_time, lianban, reason "
                "FROM zt_stocks WHERE trade_date = %s ORDER BY id", (d,))
            rows = cur.fetchall()
    finally:
        conn.close()
    out = []
    for block, code, name, last_time, lianban, reason in rows:
        if not code or not re.match(r'^\d{6}$', str(code)):
            continue
        out.append({
            '概念板块': block or '',
            '代码': str(code),
            '名称': name or '',
            '末次时间': last_time or '',
            '连扳数': lianban or '',
            '原因': reason or '',
        })
    return out


def read_daily_stats_db():
    """读取 zt_daily 涨跌家数/成交额, 返回 {YYYYMMDD: {up_count, down_count, total_amount}}"""
    try:
        conn = _dbmod.get_conn()
    except Exception as e:
        print(f'读取涨跌家数失败: {e}')
        return {}
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT to_char(trade_date,'YYYYMMDD'), up_count, down_count, total_amount FROM zt_daily")
            out = {}
            for d, u, dn, amt in cur.fetchall():
                out[d] = {
                    'up_count': u, 'down_count': dn,
                    'total_amount': float(amt) if amt is not None else None,
                }
            return out
    finally:
        conn.close()


# ============== 连扳数解析 ==============

def parse_lianban(lianban):
    """
    解析连扳数字段, 返回结构化信息

    输入示例: '1' / '2' / '3' / '1/18天9板' / '1/3天2板' / '3/7天5板'

    返回: {
        'current': int,      # 当前连板数 (斜杠前的数字, 纯数字则为其本身)
        'is_ratio': bool,    # 是否为 "X/Y天Z板" 断板回封格式
        'days': int|None,    # 周期天数 Y
        'period_count': int|None,  # 周期内涨停次数 Z
        'raw': str,
    }
    """
    raw = (lianban or '').strip()
    info = {'current': 0, 'is_ratio': False, 'days': None,
            'period_count': None, 'raw': raw}
    if not raw:
        return info
    # X/Y天Z板
    m = re.match(r'^(\d+)\s*/\s*(\d+)\s*天\s*(\d+)\s*板', raw)
    if m:
        info['current'] = int(m.group(1))
        info['is_ratio'] = True
        info['days'] = int(m.group(2))
        info['period_count'] = int(m.group(3))
        return info
    # 纯数字
    m = re.match(r'^(\d+)', raw)
    if m:
        info['current'] = int(m.group(1))
    return info


def is_925(last_time):
    """末次时间是否为 9:25 (开盘一字板)"""
    t = (last_time or '').strip().replace('：', ':')
    return t in ('9:25', '09:25', '9:25:00')


def is_kcb_cyb(code):
    """是否 300(创业板)/688(科创板) 开头"""
    return code.startswith('300') or code.startswith('688')


def make_desc(lianban, last_time):
    """
    生成个股描述
    优先级: 9:25一字板 > 1/X板格式原样 > N连板
    """
    info = parse_lianban(lianban)
    parts = []
    if is_925(last_time):
        parts.append('9:25一字板')
    if info['is_ratio']:
        parts.append(info['raw'])  # 如 1/3天2板
    elif info['current'] >= 1:
        parts.append(f"{info['current']}连板")
    return ' '.join(parts) if parts else (lianban or '')


# ============== 入选判定 ==============

def block_qualifies(stocks):
    """
    板块是否入选: 板块内任一票满足 连板>=2 / 1-X板格式 / 9:25涨停
    stocks: 该板块当日的票列表
    """
    for s in stocks:
        info = parse_lianban(s['连扳数'])
        if info['current'] >= 2:
            return True
        if info['is_ratio']:
            return True
        if is_925(s['末次时间']):
            return True
    return False


def stock_qualifies(stock):
    """
    个股是否入选(在已入选板块下):
    1/X板格式 / 9:25涨停 / 300或688开头的1板
    """
    info = parse_lianban(stock['连扳数'])
    if info['is_ratio']:
        return True
    if is_925(stock['末次时间']):
        return True
    if is_kcb_cyb(stock['代码']) and info['current'] >= 1:
        return True
    # 2连板及以上也属于值得跟踪
    if info['current'] >= 2:
        return True
    return False


def daterange_keys(all_dates, start, end):
    """返回 [start,end] 区间内、且存在Excel的日期, 升序"""
    return sorted(d for d in all_dates if start <= d <= end)


# ============== 涨幅获取 (腾讯行情接口 + 本地缓存) ==============
# 说明: 东方财富(akshare默认源)在当前网络环境直连被重置、走代理也不通,
#       改用腾讯行情接口 web.ifzq.gtimg.cn, 直连稳定可用。
#       涨跌幅 = (当日收盘 - 前一日收盘) / 前一日收盘 * 100

_PRICE_CACHE_FILE = '.price_cache.json'
_price_cache = None

import threading
_cache_lock = threading.Lock()

_TX_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
                  'AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0 Safari/537.36',
}


def _load_price_cache():
    global _price_cache
    if _price_cache is None:
        if os.path.exists(_PRICE_CACHE_FILE):
            try:
                with open(_PRICE_CACHE_FILE, 'r', encoding='utf-8') as f:
                    _price_cache = json.load(f)
            except Exception:
                _price_cache = {}
        else:
            _price_cache = {}
    return _price_cache


def _save_price_cache():
    try:
        with open(_PRICE_CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(_price_cache, f, ensure_ascii=False)
    except Exception:
        pass


# ============== 通达信 mootdx 行情 (TCP, 替代易被封的 akshare) ==============
_tdx_client = None
_tdx_lock = None
_tdx_local_ok = None  # 本地通达信可用性缓存(None=未检测)


def _get_tdx_client():
    """获取(并复用)mootdx 行情客户端。
    优先本地通达信(127.0.0.1:7709, 不限流不封IP), 不可用则回退远程标准服务器。"""
    global _tdx_client, _tdx_local_ok
    if _tdx_client is None:
        from mootdx.quotes import Quotes
        # 先尝试本地通达信客户端
        if _tdx_local_ok is None:
            try:
                _tdx_client = Quotes.factory(market='std', host='127.0.0.1', port=7709)
                # 快速验证: 拉一只票的日线, 能拿到数据说明本地客户端在线
                _df = _tdx_client.bars(symbol='000001', frequency=9, offset=1)
                _tdx_local_ok = _df is not None and len(_df) > 0
            except Exception:
                _tdx_local_ok = False
            if _tdx_local_ok:
                _ht_logger.info('使用本地通达信(127.0.0.1:7709)作为行情源')
            else:
                _ht_logger.info('本地通达信不可用, 回退远程标准服务器')
        if not _tdx_local_ok:
            _tdx_client = Quotes.factory(market='std')
    return _tdx_client


def _get_tdx_lock():
    """获取 TDX 连接的线程锁(mootdx 客户端非线程安全, 并发会协议错乱)"""
    global _tdx_lock
    if _tdx_lock is None:
        import threading
        _tdx_lock = threading.Lock()
    return _tdx_lock


def _mootdx_offset(start):
    """根据 start 到今天估算需要抓取的日线根数(含60日回看)"""
    from datetime import datetime, date
    try:
        s = datetime.strptime(start, '%Y%m%d').date()
    except Exception:
        return 250
    cal_days = (date.today() - s).days + 90  # 额外回看用于 20/60 日与 MA10
    approx = int(cal_days * 5 / 7) + 15       # 折算交易日
    return min(max(approx, 90), 800)


def fetch_range_mootdx(code, start, end):
    """
    用通达信 mootdx 获取个股日线, 计算并缓存: 当日涨跌幅 / 20日 / 60日 / MA10 / 是否跌破MA10
    涨跌幅 = (当日收盘 - 前一日收盘) / 前一日收盘 * 100  (未复权)
    返回: True 成功 / False 失败
    """
    try:
        # 北交所(8开头)mootdx std市场不支持, 直接返回False
        if code.startswith('8'):
            print(f"mootdx {code} 北交所股票, 跳过")
            return 'bse'  # 特殊返回值区分原因
        client = _get_tdx_client()
        offset = _mootdx_offset(start)
        df = client.bars(symbol=code, frequency=9, offset=offset)  # 9=日线
        if df is None or len(df) == 0:
            print(f"mootdx {code} 无数据")
            return False

        df = df.sort_index()
        closes = [float(x) for x in df['close'].tolist()]
        dates = [str(idx)[:10].replace('-', '') for idx in df.index]

        cache = _load_price_cache()
        with _cache_lock:
            # 记录该股实际有交易的日期集合, 供完整性检查区分"停牌日"(正常None)与"真缺数据"
            cache[f'{code}_tdays'] = dates
            n = len(closes)
            for i in range(n):
                dstr = dates[i]
                # 当日涨跌幅
                if i > 0 and closes[i - 1] > 0:
                    cache[f'{code}_{dstr}'] = round((closes[i] - closes[i - 1]) / closes[i - 1] * 100, 2)
                # 20 日涨幅
                if i > 0:
                    j20 = max(0, i - 19)
                    if closes[j20] > 0:
                        cache[f'{code}_{dstr}_20d'] = round((closes[i] - closes[j20]) / closes[j20] * 100, 2)
                    j60 = max(0, i - 59)
                    if closes[j60] > 0:
                        cache[f'{code}_{dstr}_60d'] = round((closes[i] - closes[j60]) / closes[j60] * 100, 2)
                # MA10 及是否跌破
                if i >= 9:
                    ma10 = sum(closes[i - 9:i + 1]) / 10
                    cache[f'{code}_{dstr}_ma10'] = round(ma10, 2)
                    cache[f'{code}_{dstr}_below_ma10'] = closes[i] < ma10
                else:
                    cache[f'{code}_{dstr}_ma10'] = None
                    cache[f'{code}_{dstr}_below_ma10'] = False
            _save_price_cache()
        return True
    except Exception as e:
        print(f"mootdx获取{code}失败: {e}")
        return False


def fetch_range_local_tdx(code, start, end):
    """
    连接本地通达信客户端(127.0.0.1:7709)拉取日线数据。
    作为 mootdx 标准服务器失败后的备用数据源。
    返回: True 成功 / False 失败
    """
    try:
        from mootdx.quotes import Quotes
        client = Quotes.factory(market='std', host='127.0.0.1', port=7709)
        offset = _mootdx_offset(start)
        df = client.bars(symbol=code, frequency=9, offset=offset)
        if df is None or len(df) == 0:
            print(f"本地通达信 {code} 无数据")
            return False

        df = df.sort_index()
        closes = [float(x) for x in df['close'].tolist()]
        dates = [str(idx)[:10].replace('-', '') for idx in df.index]

        cache = _load_price_cache()
        with _cache_lock:
            cache[f'{code}_tdays'] = dates
            n = len(closes)
            for i in range(n):
                dstr = dates[i]
                if i > 0 and closes[i - 1] > 0:
                    cache[f'{code}_{dstr}'] = round((closes[i] - closes[i - 1]) / closes[i - 1] * 100, 2)
                if i > 0:
                    j20 = max(0, i - 19)
                    if closes[j20] > 0:
                        cache[f'{code}_{dstr}_20d'] = round((closes[i] - closes[j20]) / closes[j20] * 100, 2)
                    j60 = max(0, i - 59)
                    if closes[j60] > 0:
                        cache[f'{code}_{dstr}_60d'] = round((closes[i] - closes[j60]) / closes[j60] * 100, 2)
                if i >= 9:
                    ma10 = sum(closes[i - 9:i + 1]) / 10
                    cache[f'{code}_{dstr}_ma10'] = round(ma10, 2)
                    cache[f'{code}_{dstr}_below_ma10'] = closes[i] < ma10
                else:
                    cache[f'{code}_{dstr}_ma10'] = None
                    cache[f'{code}_{dstr}_below_ma10'] = False
            _save_price_cache()
        return True
    except Exception as e:
        print(f"本地通达信获取{code}失败: {e}")
        return False


def fetch_range(code, start, end):
    """获取个股涨幅数据: 优先 mootdx 远程 -> 失败则尝试本地通达信 -> 仍失败返回 False"""
    _ht_logger.debug(f'fetch_range {code} {start}~{end}')
    result = fetch_range_mootdx(code, start, end)
    if result is True:
        _ht_logger.debug(f'fetch_range {code} mootdx成功')
        return True
    if result == 'bse':
        _ht_logger.debug(f'fetch_range {code} 北交所跳过')
        return False
    _ht_logger.debug(f'fetch_range {code} mootdx失败, 尝试本地通达信')
    print(f"{code} mootdx远程失败, 尝试本地通达信...")
    if fetch_range_local_tdx(code, start, end):
        _ht_logger.debug(f'fetch_range {code} 本地通达信成功')
        return True
    _ht_logger.warning(f'fetch_range {code} 全部失败')
    print(f"{code} 本地通达信也失败, 无法获取数据")
    return False


def get_pct_change(code, date):
    """从缓存读取某日涨跌幅, 未命中返回 None"""
    cache = _load_price_cache()
    return cache.get(f'{code}_{date}')


def get_pct_20d(code, date):
    """从缓存读取某日20日涨幅, 未命中返回 None"""
    cache = _load_price_cache()
    return cache.get(f'{code}_{date}_20d')


def get_pct_60d(code, date):
    """从缓存读取某日60日涨幅, 未命中返回 None"""
    cache = _load_price_cache()
    return cache.get(f'{code}_{date}_60d')


def get_ma10(code, date):
    """从缓存读取某日10日均线, 未命中返回 None"""
    cache = _load_price_cache()
    return cache.get(f'{code}_{date}_ma10')


def is_below_ma10(code, date):
    """判断某日是否跌破10日线, 未命中返回 None"""
    cache = _load_price_cache()
    return cache.get(f'{code}_{date}_below_ma10')


def is_suspended(code, date):
    """判断某股在某日是否停牌: 即该日不在 mootdx 返回的实际交易日列表中。
    若无 _tdays 缓存(从未成功拉取过), 返回 None 表示未知。"""
    cache = _load_price_cache()
    tdays = cache.get(f'{code}_tdays')
    if tdays is None:
        return None
    return date not in tdays


def prefetch_prices(codes, start, end, progress=None, check_dates=None):
    """
    批量预取涨幅(通达信 mootdx)。
    check_dates: 实际交易日列表(YYYYMMDD)。提供时按逐日核对完整性——
                 只要任一交易日缺 pct 或 MA10 就重取(mootdx 一次拉全, 能补齐个别缺失日)。
                 不提供则回退到工作日估算。
    """
    _report = progress if callable(progress) else (lambda *a, **k: None)
    cache = _load_price_cache()

    if check_dates:
        query_dates = list(check_dates)
    else:
        # 回退: 生成查询范围内的工作日
        from datetime import datetime, timedelta
        try:
            s_dt = datetime.strptime(start, '%Y%m%d')
            e_dt = datetime.strptime(end, '%Y%m%d')
        except Exception:
            return
        query_dates = []
        current = s_dt
        while current <= e_dt:
            if current.weekday() < 5:
                query_dates.append(current.strftime('%Y%m%d'))
            current += timedelta(days=1)

    # 逐个交易日核对: 任一交易日缺 pct 或 MA10 即需要重取
    # 注意: 已标记为无数据(cache['{c}_no_data']=True)的票跳过重拉, 但仍加入 failed 列表
    need_fetch = []
    known_no_data = []  # 上次已确认无数据的票
    _ht_logger.debug(f'prefetch_prices 共 {len(codes)} 只, 日期 {start}~{end}')
    for c in codes:
        if cache.get(f'{c}_no_data'):
            known_no_data.append({'code': c, 'reason': cache.get(f'{c}_no_data_reason', '无数据')})
            continue
        incomplete = any((f'{c}_{d}' not in cache) or (f'{c}_{d}_ma10' not in cache)
                         for d in query_dates)
        if incomplete:
            need_fetch.append(c)
    
    total_codes = len(codes)
    cached_cnt = total_codes - len(need_fetch)
    if not need_fetch:
        _ht_logger.debug(f'prefetch_prices 全部命中缓存 {total_codes} 只')
        print(f"所有 {total_codes} 只股票涨幅数据已缓存")
        _report('price', f'行情已全部缓存 ({total_codes} 只)', total_codes, total_codes)
        all_failed = known_no_data
        if all_failed:
            print(f"以下 {len(all_failed)} 只股票无法获取行情数据(将显示为空):")
            for f in all_failed:
                print(f"  {f['code']} - {f['reason']}")
        return {'success': total_codes - len(known_no_data), 'failed': all_failed, 'total': total_codes}

    _ht_logger.debug(f'prefetch_prices 需拉取 {len(need_fetch)} 只, 已缓存 {cached_cnt} 只, known_no_data {len(known_no_data)} 只')
    print(f"需要获取 {len(need_fetch)} 只股票的涨幅数据...")
    import time
    success_list = []
    failed_list = []

    for i, c in enumerate(need_fetch):
        try:
            ok = fetch_range(c, start, end)
        except Exception as e:
            print(f"获取 {c} 失败: {e}")
            ok = False
            failed_list.append({'code': c, 'reason': str(e)})
            time.sleep(0.2)
        else:
            time.sleep(0.05)
        if ok:
            success_list.append(c)
            # fetch 成功但 mootdx 可能未覆盖所有交易日(停牌/未上市/数据源缺失)
            # 对仍缺的日期写 None 占位, 避免下次因 key 不存在而反复重取
            _cache = _load_price_cache()
            with _cache_lock:
                _filled = False
                for _d in query_dates:
                    for _sfx in ('', '_ma10', '_below_ma10', '_20d', '_60d'):
                        k = f'{c}_{_d}{_sfx}'
                        if k not in _cache:
                            _cache[k] = None
                            _filled = True
                if _filled:
                    _save_price_cache()
        else:
            # 拉取失败: 写 _no_data 标记避免完整性检查反复重拉, 同时保留在 failed 列表中
            reason = next((f['reason'] for f in failed_list if f['code'] == c), '无数据返回')
            _cache = _load_price_cache()
            with _cache_lock:
                _cache[f'{c}_no_data'] = True
                _cache[f'{c}_no_data_reason'] = reason
                # 同时写 None 占位避免其他地方的 key-not-in 检查
                for _d in query_dates:
                    for _sfx in ('', '_ma10', '_below_ma10', '_20d', '_60d'):
                        k = f'{c}_{_d}{_sfx}'
                        if k not in _cache:
                            _cache[k] = None
                _save_price_cache()
            if not any(f['code'] == c for f in failed_list):
                failed_list.append({'code': c, 'reason': reason})
        _report('price', f'获取行情 {i + 1}/{len(need_fetch)} (代码 {c})',
                cached_cnt + i + 1, total_codes)
        if (i + 1) % 20 == 0:
            print(f"已处理 {i + 1}/{len(need_fetch)}，成功 {len(success_list)}")
    
    print(f"完成: 成功获取 {len(success_list)}/{len(need_fetch)} 只股票数据")
    all_failed = known_no_data + failed_list
    if all_failed:
        print(f"以下 {len(all_failed)} 只股票无法获取行情数据(将显示为空):")
        for f in all_failed:
            print(f"  {f['code']} - {f['reason']}")

    return {
        'success': len(success_list),
        'failed': all_failed,
        'total': len(need_fetch),
        'cached': len(codes) - len(need_fetch) - len(known_no_data)
    }


def apply_removal_rules(data, progress=None, manual_remove_codes=None):
    """
    对已有的 track_hot_stocks 结果应用移除规则。
    直接在内存中操作, 不重新读库/拉行情。
    manual_remove_codes: 可选, 需手动剔除(不再追踪)的股票代码集合/列表。
                         这些票会被标记为在所有日期都不显示(仅本次生效, 不写缓存)。
    返回: 修改后的 data（原地修改 by_date 中的 stocks 列表）
    """
    _report = progress if callable(progress) else (lambda *a, **k: None)
    dates = data.get('dates', [])
    date_pos = {dd: i for i, dd in enumerate(dates)}
    by_date = data.get('by_date', [])
    manual_set = set(manual_remove_codes) if manual_remove_codes else set()

    # 重建 block_cum_stocks 结构（从 by_date 反推）
    block_cum = {}  # block -> {code: stock_item}
    for day in by_date:
        for b in day.get('blocks', []):
            blk = b['block']
            if blk not in block_cum:
                block_cum[blk] = {}
            for s in b.get('stocks', []):
                if s['code'] not in block_cum[blk]:
                    block_cum[blk][s['code']] = s

    # 应用移除规则
    for blk, stocks_map in block_cum.items():
        for code, st in stocks_map.items():
            st.pop('warn_date', None)
            st.pop('remove_date', None)
            st.pop('remove_reason', None)

    # 手动移除: 标记一个早于所有交易日的 remove_date, 使其在任何日期都不显示
    if manual_set:
        _manual_marker = '00000000'
        for stocks_map in block_cum.values():
            for code, st in stocks_map.items():
                if code in manual_set:
                    st['remove_date'] = _manual_marker
                    st['remove_reason'] = '手动移除(无法获取行情)'

    for d in dates:
        di = date_pos.get(d)
        for blk, stocks_map in block_cum.items():
            for code, st in stocks_map.items():
                if st.get('remove_date'):
                    continue
                # 只从个股入选日(first_date)起才开始检查跌破10日线
                # 否则会在个股尚未入选时就因历史价格数据触发预警/移除, 导致漏票
                if st.get('first_date', '') > d:
                    continue
                track = st.get('track', {}).get(d)
                if not track:
                    continue
                pct = track.get('pct')
                # 停牌日不参与跌破判断, 也不清除已有预警状态
                if track.get('suspended'):
                    continue
                is_kcb_cyb = st.get('is_kcb_cyb', False)
                is_limit_up = False
                if pct is not None:
                    is_limit_up = pct >= 19.4 if is_kcb_cyb else pct >= 9.8
                if '一字板' in (track.get('desc', '') or ''):
                    is_limit_up = True
                below = (track.get('below_ma10') == True) and not is_limit_up
                if below:
                    if st.get('warn_date') is None:
                        st['warn_date'] = d
                    wi = date_pos.get(st['warn_date'])
                    if wi is not None and di is not None and (di - wi + 1) >= 3:
                        st['remove_date'] = d
                        st['remove_reason'] = '跌破10日线第三日未收回'
                else:
                    st['warn_date'] = None
        _report('remove', f'{d}: 移除规则处理完成', dates.index(d) + 1, len(dates))

    # 更新 by_date 中每天的 stocks 列表(过滤已移除的票)
    # 必须以 block_cum 为准遍历所有累积票，而不是 b['stocks']（后者可能只是当天子集）
    for day in by_date:
        d = day['date']
        for b in day.get('blocks', []):
            blk = b['block']
            active, removed_today = [], []
            for st in block_cum.get(blk, {}).values():
                # 只显示 first_date <= 当天 的票（还未入选的不显示）
                if st.get('first_date', '') > d:
                    continue
                remove_date = st.get('remove_date')
                if remove_date is None or d < remove_date:
                    active.append(st)
                elif remove_date == d:
                    active.append(st)
                    removed_today.append(st['code'])
            b['stocks'] = active
            b['active_count'] = len(active)
            b['removed_today'] = removed_today

    return data


def track_hot_stocks(start, end, sort='stock_count', with_price=True,
                     excel_dir=EXCEL_DIR, progress=None, source='db',
                     apply_removal=True):
    """
    热门股追踪主函数

    参数:
        start, end: 8位日期字符串 YYYYMMDD
        sort: 'stock_count'(按板块内入选票数) | 'days'(按板块出现天数)
        with_price: 是否获取涨幅(mootdx)

    返回: dict
        {
          'start','end','sort',
          'dates': [日期...],            # 区间内有数据的交易日
          'by_date': [                   # 按日期分组
            {
              'date': '20260601',
              'blocks': [                # 当日"新加入"的板块
                {
                  'block': 板块名,
                  'new_count': 当日新增入选票数,
                  'total_count': 板块累计入选票数,
                  'stocks': [
                    {'code','name','is_kcb_cyb','first_date',
                     'track': {'20260601': {'desc','pct','present'}, ...}}
                  ]
                }
              ]
            }
          ]
        }
    排序由前端按 new_count / total_count 实时切换, 不需重新请求
    """
    _report = progress if callable(progress) else (lambda *a, **k: None)

    # 数据源: db(默认, 读 PostgreSQL) 或 excel
    if source == 'db':
        all_files = None
        dates = sorted(d for d in list_db_dates() if start <= d <= end)
    else:
        all_files = list_excel_dates(excel_dir)
        dates = daterange_keys(all_files.keys(), start, end)

    # === 阶段1: 逐日载入数据 + 按入榜规则加入板块/个股 ===
    _report('build', f'开始载入数据 (来源: {"数据库" if source == "db" else "Excel"})，共 {len(dates)} 个交易日', 0, len(dates))
    daily = {}      # date -> block -> list[stock] (已排除市场连板股)
    daily_all = {}  # date -> 原始所有个股(含市场连板股), 用于左侧高位股统计
    for i, d in enumerate(dates):
        rows = read_day_stocks_db(d) if source == 'db' else parse_excel_04(all_files[d])
        daily_all[d] = rows
        bmap = {}
        for s in rows:
            blk = s['概念板块']
            if blk in EXCLUDED_BLOCKS:
                continue
            bmap.setdefault(blk, []).append(s)
        daily[d] = bmap
        _report('build', f'{d}: 载入 {len(rows)} 只个股 / {len(bmap)} 个板块', i + 1, len(dates))

    # === 板块入选: 逐日扫描, 首次满足即加入(保留) ===
    selected_blocks = {}   # block -> first_date
    for d in dates:
        for blk, stocks in daily[d].items():
            if blk in selected_blocks:
                continue
            if block_qualifies(stocks):
                selected_blocks[blk] = d
    _report('build', f'入榜板块共 {len(selected_blocks)} 个', len(dates), len(dates))

    # === 个股入选 + 逐日跟踪 ===
    # 先收集所有入选板块的入选个股代码, 批量预取涨幅
    fetch_report = None
    if with_price:
        all_codes = set()
        for blk in selected_blocks:
            for d in dates:
                for s in daily[d].get(blk, []):
                    if stock_qualifies(s):
                        all_codes.add(s['代码'])
        # 左侧列需要3板+高位股(及其次日断板)的涨幅, 一并预取
        for d in dates:
            for r in daily_all[d]:
                if parse_lianban(r['连扳数']).get('current', 0) >= 3:
                    all_codes.add(r['代码'])
        
        fetch_report = None
        if all_codes and dates:
            _report('price', f'阶段2: 获取 {len(all_codes)} 只个股行情(通达信 mootdx)…', 0, len(all_codes))
            fetch_report = prefetch_prices(all_codes, dates[0], dates[-1],
                                           progress=_report, check_dates=dates)
        else:
            _report('price', '阶段2: 无需获取行情', 0, 0)

    def build_track(blk, code):
        """构建某票在某板块下逐日跟踪"""
        track = {}
        for d in dates:
            rec = None
            for s in daily[d].get(blk, []):
                if s['代码'] == code:
                    rec = s
                    break
            pct = get_pct_change(code, d) if with_price else None
            pct_20d = get_pct_20d(code, d) if with_price else None
            pct_60d = get_pct_60d(code, d) if with_price else None
            below_ma10 = is_below_ma10(code, d) if with_price else None
            ma10 = get_ma10(code, d) if with_price else None
            suspended = is_suspended(code, d) if with_price else None
            if rec:
                track[d] = {
                    'desc': make_desc(rec['连扳数'], rec['末次时间']),
                    'lianban': rec['连扳数'],
                    'time': rec['末次时间'],
                    'reason': rec['原因'],
                    'pct': pct,
                    'pct_20d': pct_20d,
                    'pct_60d': pct_60d,
                    'ma10': ma10,
                    'below_ma10': below_ma10,
                    'suspended': suspended,
                    'present': True,
                }
            else:
                track[d] = {
                    'present': False, 
                    'pct': pct,
                    'pct_20d': pct_20d,
                    'pct_60d': pct_60d,
                    'ma10': ma10,
                    'below_ma10': below_ma10,
                    'suspended': suspended,
                }
        return track

    # === 按日期分组: 每个日期显示截至当天累计入选的板块及其完整阵容 ===
    # 全局指标(整个区间, 用于一致排序):
    #   block_total: 累计入选票数
    #   block_times: 累计出现次数(区间内该板块出现的天数)
    block_total = {}      # block -> 累计入选票数
    block_times = {}      # block -> 累计出现天数
    for blk in selected_blocks:
        codes = set()
        for d in dates:
            for s in daily[d].get(blk, []):
                if stock_qualifies(s):
                    codes.add(s['代码'])
        block_total[blk] = len(codes)
        block_times[blk] = sum(1 for d in dates if blk in daily[d])

    _report('remove', '阶段3: 应用移除规则(跌破10日线预警, 第三日仍破位则移除)…', 0, len(dates))
    date_pos = {dd: i for i, dd in enumerate(dates)}  # 交易日 -> 序号
    by_date = []
    # 累积每个板块已入选的票(跨日累加), 实现"每天看完整阵容"
    block_cum_stocks = {}   # block -> list[stock_item] (按入选顺序)
    block_seen_codes = {}   # block -> set(code)
    block_removed_stocks = {}  # block -> {code: remove_date} 被移除的股票及其移除日期
    for blk in selected_blocks:
        block_cum_stocks[blk] = []
        block_seen_codes[blk] = set()
        block_removed_stocks[blk] = {}
    block_is_dormant = {blk: False for blk in selected_blocks}  # 板块是否休眠(所有票已移除)

    for d in dates:
        day_blocks = []
        for blk in selected_blocks:
            blk_stocks = daily[d].get(blk, [])
            is_dormant = block_is_dormant[blk]

            if is_dormant:
                # 休眠板块: 需 block_qualifies(连板>=2 / X/Y天Z板 / 9:25) 才能重新唤醒
                candidates = [s for s in blk_stocks if s['代码'] not in block_seen_codes[blk]]
                if candidates and block_qualifies(candidates):
                    # 唤醒成功: 加入所有满足 stock_qualifies 的票
                    for s in candidates:
                        if stock_qualifies(s):
                            code = s['代码']
                            block_seen_codes[blk].add(code)
                            block_cum_stocks[blk].append({
                                'code': code,
                                'name': s['名称'],
                                'is_kcb_cyb': is_kcb_cyb(code),
                                'first_date': d,
                                'track': build_track(blk, code),
                            })
            else:
                # 活跃板块: 按 stock_qualifies 正常加入新票
                for s in blk_stocks:
                    code = s['代码']
                    if code in block_seen_codes[blk]:
                        continue
                    if stock_qualifies(s):
                        block_seen_codes[blk].add(code)
                        block_cum_stocks[blk].append({
                            'code': code,
                            'name': s['名称'],
                            'is_kcb_cyb': is_kcb_cyb(code),
                            'first_date': d,
                            'track': build_track(blk, code),
                        })
            
            cum = block_cum_stocks[blk]
            
            # 移除规则(新): 跌破10日线当日"预警", 自预警日起第三日收盘仍<10日线则移除;
            # 期间收回10日线(或涨停)则解除预警。apply_removal=False 时不做(数据缺失阶段)。
            for st in (cum if apply_removal else []):
                if st.get('remove_date'):
                    continue  # 已移除, 不再处理
                track = st['track'].get(d)
                if not track:
                    continue
                # 涨停判定(涨停/一字板不算破位)
                pct = track.get('pct')
                is_limit_up = False
                if pct is not None:
                    is_limit_up = pct >= 19.4 if st['is_kcb_cyb'] else pct >= 9.8
                if '一字板' in (track.get('desc', '') or ''):
                    is_limit_up = True

                below = (track.get('below_ma10') == True) and not is_limit_up

                if below:
                    # 首次跌破: 进入预警
                    if st.get('warn_date') is None:
                        st['warn_date'] = d
                    # 自预警日起的交易日序数(预警日记为第1日)
                    wi = date_pos.get(st['warn_date'])
                    di = date_pos.get(d)
                    if wi is not None and di is not None and (di - wi + 1) >= 3:
                        # 第三日仍收盘破位 -> 移除
                        st['remove_date'] = d
                        st['remove_reason'] = '跌破10日线第三日未收回'
                        block_removed_stocks[blk][st['code']] = d
                else:
                    # 收回10日线或涨停, 解除预警
                    st['warn_date'] = None
            
            # 计算当前仍在跟踪的股票（未被移除或在移除当天仍显示）
            active_stocks = []
            removed_today = []
            for st in cum:
                remove_date = st.get('remove_date')
                if remove_date is None:
                    # 未被移除，继续跟踪
                    active_stocks.append(st)
                elif remove_date == d:
                    # 当天移除，仍然显示但标记为已移除
                    active_stocks.append(st)
                    removed_today.append(st['code'])
                # else: 已移除且不是当天，不再显示
            
            # 计算板块是否即将移除（所有股票都已移除）
            all_removed = len(cum) > 0 and len(active_stocks) == len([st for st in cum if st.get('remove_date')])
            block_removing = all_removed and len(removed_today) > 0
            # 更新板块休眠状态: 所有票已移除且不再显示 → 进入休眠
            block_is_dormant[blk] = (len(active_stocks) == 0 and len(cum) > 0)
            
            # 任何入选过的板块每天都显示(含累积阵容为空的早期日期), 保证列纵向对齐
            new_today = sum(1 for st in cum if st['first_date'] == d)
            day_blocks.append({
                'block': blk,
                'cum_count': len(cum),               # 截至当天累计入选票数
                'active_count': len(active_stocks),  # 当前仍在跟踪的票数
                'new_count': new_today,              # 当天新入选票数
                'removed_today': removed_today,      # 当天移除的股票代码列表
                'times_count': block_times[blk],     # 板块累计出现次数(全局)
                'total_count': block_total[blk],     # 板块累计票数(全局)
                'active': blk in daily[d],           # 当天该板块是否在复盘中出现
                'removing': block_removing,          # 板块是否即将移除
                'stocks': [dict(st) for st in active_stocks],
            })
        if day_blocks:
            by_date.append({'date': d, 'blocks': day_blocks})
        _report('remove', f'{d}: 移除规则处理完成', dates.index(d) + 1, len(dates))

    # === 板块图表汇总数据 ===
    # 每个板块: 逐日平均涨幅 / 逐日累计平均涨幅 / 逐日累计入选票数 / 各入选票逐日涨幅序列
    blocks_summary = []
    for blk in selected_blocks:
        roster = block_cum_stocks[blk]
        if not roster:
            continue
        avg_series = {}          # 单日平均涨幅
        cum_avg_series = {}      # 累计平均涨幅
        cum_series = {}
        for d in dates:
            # 截至当天累计入选的票
            cum_stocks = [st for st in roster if st['first_date'] <= d]
            cum_series[d] = len(cum_stocks)
            # 单日平均涨幅
            pcts = [st['track'][d].get('pct') for st in cum_stocks
                    if isinstance(st['track'][d].get('pct'), (int, float))]
            avg_series[d] = round(sum(pcts) / len(pcts), 2) if pcts else None
        # 计算累计平均涨幅
        cum_sum = 0
        cum_count = 0
        for d in dates:
            if avg_series[d] is not None:
                cum_sum += avg_series[d]
                cum_count += 1
            cum_avg_series[d] = round(cum_sum, 2) if cum_count > 0 else None
        stocks_series = [{
            'code': st['code'],
            'name': st['name'],
            'first_date': st['first_date'],
            'pct': {d: st['track'][d].get('pct') for d in dates},
        } for st in roster]
        blocks_summary.append({
            'block': blk,
            'times_count': block_times[blk],
            'total_count': block_total[blk],
            'avg_pct': cum_avg_series,   # 改为累计平均涨幅
            'daily_avg_pct': avg_series, # 单日平均涨幅（保留）
            'cum_count': cum_series,
            'stocks': stocks_series,
        })

    # === 左侧汇总列: 每日涨跌数 + 3板及以上高位股 + 昨日高位今日断板 ===
    daily_stats = read_daily_stats_db()
    date_summary = {}
    prev_high = None  # 上一交易日的3板+个股(用于断板追踪)
    for d in dates:
        # 汇总当日每只票(取最大连板; 概念优先非市场连板股)
        info = {}
        for r in daily_all.get(d, []):
            code = r['代码']
            cur = parse_lianban(r['连扳数']).get('current', 0)
            blk = r['概念板块']
            e = info.get(code)
            if e is None:
                info[code] = {'code': code, 'name': r['名称'], 'lianban': r['连扳数'],
                              'current': cur, 'block': blk, 'reason': r.get('原因', '')}
            else:
                if cur > e['current']:
                    e['current'] = cur
                    e['lianban'] = r['连扳数']
                if blk != '市场连板股' and e['block'] == '市场连板股':
                    e['block'] = blk
                    e['reason'] = r.get('原因', e.get('reason', ''))
        high = sorted([v for v in info.values() if v['current'] >= 3],
                      key=lambda x: -x['current'])
        # 断板: 昨日3板+今日不在涨停列表(未涨停)
        broken = []
        if prev_high is not None:
            today_codes = set(info.keys())
            for pc in prev_high:
                if pc['code'] not in today_codes:
                    pct = get_pct_change(pc['code'], d) if with_price else None
                    broken.append({'code': pc['code'], 'name': pc['name'],
                                   'block': pc['block'], 'pct': pct,
                                   'lianban': pc.get('lianban', ''),
                                   'current': pc.get('current', 0),
                                   'reason': pc.get('reason', '')})
        stats = daily_stats.get(d, {})
        date_summary[d] = {
            'up_count': stats.get('up_count'),
            'down_count': stats.get('down_count'),
            'total_amount': stats.get('total_amount'),
            'high_stocks': [{'code': v['code'], 'name': v['name'],
                             'lianban': v['lianban'], 'block': v['block'],
                             'reason': v.get('reason', '')} for v in high],
            'broken_stocks': broken,
        }
        prev_high = high

    _report('done', f'计算完成: {len(selected_blocks)} 个板块, {len(dates)} 个交易日', len(dates), len(dates))
    return {
        'start': start,
        'end': end,
        'sort': sort,
        'dates': dates,
        'by_date': by_date,
        'blocks_summary': blocks_summary,
        'date_summary': date_summary,
        'fetch_report': fetch_report,
    }


if __name__ == '__main__':
    # 命令行自测: py hot_track.py 20260601 20260605
    start = sys.argv[1] if len(sys.argv) > 1 else '20260601'
    end = sys.argv[2] if len(sys.argv) > 2 else '20260605'
    data = track_hot_stocks(start, end, with_price=False)
    print(f"日期范围: {start}~{end}, 有数据日: {data['dates']}\n")
    for day in data['by_date']:
        print(f"【{day['date']}】 新增板块 {len(day['blocks'])} 个")
        for b in day['blocks']:
            print(f"  [{b['block']}] 当日新增{b['new_count']}票 累计{b['total_count']}票")
            for s in b['stocks']:
                seq = [f"{d[4:]}:{s['track'][d]['desc']}"
                       for d in data['dates'] if s['track'][d]['present']]
                print(f"      {s['name']}({s['code']}) {' | '.join(seq)}")
        print()
