#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
涨停复盘数据提取（GLM-OCR + 红色板块标题识别）

流程（找图与提取合并，直接从 _order.txt 第3/4张开始，不轮询）:
- 03图 = _order.txt 第3张: OCR 后开头200字含 "880005" 或 "涨跌家数"
        -> 提取涨跌家数/成交额，否则判定03不存在
- 04图 = _order.txt 第4张: extract_sectors_by_red 识别的板块标题含 "市场连板股"
        -> 提取板块个股，否则判定04不存在
- 04图板块归属: 红色检测定位板块标题 -> 裁剪单条小图用视觉模型识别板块名
  -> 按 y 坐标把 OCR 提取的股票分配到对应板块（涨停炸板之后的数据丢弃）
- 只要 03/04 命中其一即生成 Excel，缺失方在表内写明原因；都不存在则失败
- 提取数据并按模板导出 Excel 到 excelDataSource

用法:
    py extract_glm.py 20260629            # 提取单日
    py extract_glm.py 20260629 20260626   # 提取多日
    py extract_glm.py month 202606        # 按月批量(跳过已导出)
    py extract_glm.py all                 # 全部日期批量
"""

# 启动时自动检测依赖
from bootstrap import ensure_dependencies
ensure_dependencies({
    'requests': 'requests>=2.31.0',
    'bs4': 'beautifulsoup4>=4.12.0',
    'yaml': 'PyYAML>=6.0',
    'openpyxl': 'openpyxl>=3.1.0',
    'PIL': 'Pillow>=10.0.0',
    'numpy': 'numpy>=1.24.0',
})

import os
import re
import json
import base64
import sys
import time

# Windows 控制台默认 GBK，强制 stdout/stderr 为 UTF-8，避免 ✓/✗ 等字符报错
try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

import yaml
import requests
from bs4 import BeautifulSoup
import openpyxl


# ============== 配置 ==============

def load_config():
    config_path = 'config.yml'
    if os.path.exists(config_path):
        with open(config_path, 'r', encoding='utf-8') as f:
            return yaml.safe_load(f)
    return {}


CONFIG = load_config()

GLM_OCR_URL = "https://open.bigmodel.cn/api/paas/v4/layout_parsing"
GLM_CHAT_URL = "https://open.bigmodel.cn/api/paas/v4/chat/completions"

# 排除的板块(精确匹配); 另外任何含"其他"的板块名(如【其他热点】【其他题材】)一律排除
EXCLUDED_SECTORS = ['涨停炸板', '首板']

# 页面大标题（非板块，不吸收任何股票，强制数量0）
PAGE_TITLES = ['湖南人涨停复盘', '涨停复盘']


def _is_excluded_sector(name):
    """板块是否应排除(不写入结果, 也不计入"未分配"):
    命中精确排除项, 或板块名中包含"其他"(覆盖【其他热点】【其他个股】【其他XX】等)。"""
    if not name:
        return False
    return name in EXCLUDED_SECTORS or '其他' in name

# 提取模板路径
TEMPLATE_PATH = 'systemOriginalData/提取模板.xlsx'


def get_api_key():
    return CONFIG.get('ai', {}).get('zhipu', {}).get('api_key', '')


def get_vision_model():
    return CONFIG.get('ai', {}).get('zhipu', {}).get('vision_model', 'GLM-4.5V')


# ============== GLM-OCR 调用（带缓存） ==============

def glm_ocr(image_path, use_cache=True):
    """
    调用 GLM-OCR 识别图片，返回 layout_details（带磁盘缓存避免重复花token）

    返回: (layout_details: list, error: str|None)
    """
    cache_path = image_path + '.glmocr.json'

    # 命中缓存
    if use_cache and os.path.exists(cache_path):
        try:
            if os.path.getmtime(cache_path) >= os.path.getmtime(image_path):
                with open(cache_path, 'r', encoding='utf-8') as f:
                    return json.load(f), None
        except Exception:
            pass

    api_key = get_api_key()
    if not api_key or api_key.startswith('在这里') or api_key == 'YOUR_API_KEY_HERE':
        return None, "未配置有效的 API Key（请在 config.yml 填入 ai.zhipu.api_key）"

    with open(image_path, 'rb') as f:
        img_base64 = base64.b64encode(f.read()).decode()

    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }
    payload = {
        "model": "glm-ocr",
        "file": f"data:image/png;base64,{img_base64}"
    }

    # 带重试的请求（GLM-OCR 偶发 SSL/超时）
    result = None
    last_err = None
    for attempt in range(1, 4):
        try:
            resp = requests.post(GLM_OCR_URL, headers=headers, json=payload, timeout=180)
            result = resp.json()
            last_err = None
            break
        except Exception as e:
            last_err = e
            if attempt < 3:
                wait = 2 ** (attempt - 1)
                print(f"    GLM-OCR 网络异常，{wait}s 后重试 [{attempt}/2]...")
                time.sleep(wait)
    if last_err is not None:
        return None, f"请求异常: {last_err}"

    if 'layout_details' not in result:
        return None, result.get('error', {}).get('message', f'未知错误: {result}')

    layout_details = result['layout_details']

    # 写缓存
    try:
        with open(cache_path, 'w', encoding='utf-8') as f:
            json.dump(layout_details, f, ensure_ascii=False)
    except Exception:
        pass

    return layout_details, None


# ============== 文本/表格解析 ==============

def layout_to_text(layout_details, limit=None):
    """把 layout_details 按顺序拼成纯文本（用于关键字判断）"""
    parts = []
    for page in layout_details:
        for item in page:
            content = item.get('content', '')
            if not content:
                continue
            if item.get('label') == 'table':
                soup = BeautifulSoup(content, 'html.parser')
                parts.append(soup.get_text(separator=' '))
            else:
                parts.append(content)
    text = '\n'.join(parts)
    return text[:limit] if limit else text


def parse_html_tables(layout_details):
    """解析 GLM-OCR 返回的 HTML 表格 -> [[row, ...], ...]"""
    all_tables = []
    for page in layout_details:
        for item in page:
            if item.get('label') == 'table':
                soup = BeautifulSoup(item.get('content', ''), 'html.parser')
                for table in soup.find_all('table'):
                    table_data = []
                    for row in table.find_all('tr'):
                        cells = row.find_all(['th', 'td'])
                        row_data = [c.get_text(strip=True) for c in cells]
                        if row_data:
                            table_data.append(row_data)
                    if table_data:
                        all_tables.append(table_data)
    return all_tables


def extract_zhangdie(layout_details):
    """
    从03图提取涨跌数据
    返回: {'上涨家数', '下跌家数', '总成交额'}
    总成交额: 匹配"总成交额 X"后的数字(单位可能是亿/万, OCR偶发误识,
    因数值量级一致(均为万亿级), 直接取数值不换算)
    """
    result = {'上涨家数': None, '下跌家数': None, '总成交额': None}
    text = layout_to_text(layout_details)

    m = re.search(r'上涨家数\s*(\d+)', text)
    if m:
        result['上涨家数'] = int(m.group(1))
    m = re.search(r'下跌家数\s*(\d+)', text)
    if m:
        result['下跌家数'] = int(m.group(1))
    # 总成交额: 取"总成交额"后的数值(兼容 亿/万/(亿)/(万) 等单位, 不做换算)
    m = re.search(r'总成交额\s*([\d.]+)', text)
    if m:
        result['总成交额'] = m.group(1)

    return result


# 板块标题行: 【板块名】N只 X亿  或  【板块名】
_SECTOR_HEAD_RE = re.compile(r'【([^】]+)】')


def _iter_blocks(layout_details):
    """按原始顺序产出 (label, content) 块"""
    for page in layout_details:
        for item in page:
            yield item.get('label', ''), item.get('content', '')


# 分段参数
SEG_HEIGHT = 4000      # 每段高度(px) - 增大以减少分段数
SEG_OVERLAP = 500      # 段间重叠(px)，避免板块被腰斩
SEG_TRIGGER = 4500     # 超过此高度才分段，否则整图一次


def ensure_dependencies_extra():
    """确保 numpy 可用（红色检测需要）"""
    ensure_dependencies({'numpy': 'numpy>=1.24.0'})


# 板块标题单条识别 prompt
_TITLE_PROMPT = """这是从股票复盘图裁剪的一小条，最上方有一行红色文字的板块标题，格式为【板块名】N只 X亿。
请输出最上方那一行红色板块标题中的板块名称（去掉【】符号）和"N只"中的数字N（该板块的个股数量）。
只输出一个JSON对象，例如：{"板块":"半导体","数量":5}
如果看不到红色板块标题，输出：{}"""


def _scan_red_rows(arr, threshold, gap, min_height):
    """在给定阈值下扫描红色行并合并成区间, 返回 [(y0, y1), ...]。"""
    import numpy as np
    R = arr[:, :, 0].astype(int)
    G = arr[:, :, 1].astype(int)
    B = arr[:, :, 2].astype(int)
    red_mask = (R > 140) & (G < 100) & (B < 100)
    rows = np.where(red_mask.sum(axis=1) >= threshold)[0]
    bands = []
    if len(rows) > 0:
        s = p = rows[0]
        for r in rows[1:]:
            if r - p > gap:
                if p - s >= min_height:  # 过滤过矮的噪点区间
                    bands.append((int(s), int(p)))
                s = r
            p = r
        if p - s >= min_height:
            bands.append((int(s), int(p)))
    return bands


def _find_red_bands(arr, threshold=None, gap=15, min_height=8, max_bands=30):
    """
    检测图片中红色文字的 y 坐标区间（板块标题为红色）

    参数:
        arr: numpy RGB 数组
        threshold: 一行最少红色像素数。None=自适应(见下)。
        gap: 区间合并间隔(px)
        min_height: 最小区间高度(px)，过滤噪点
        max_bands: 自适应模式下允许的最多标题数, 超出则提高阈值重扫

    自适应说明:
        复盘图里上涨个股的名称/价格也是红色。阈值过低(如5)会把零散红色价格
        文字误判成标题, 一张图检出上百个"板块", 触发上百次 Vision 调用而看似卡死;
        阈值过高(如宽度的6%)又会漏掉真实的窄板块标题(如 PCB/电力/锂电池), 导致
        这些板块下的个股无处归属。经实测, 真实标题在"宽度约3%"这一档能抓全,
        而噪点图在同档也只有十来个区间。故默认从较低阈值起步抓全标题, 仅当某图
        异常地检出过多区间(疑似噪点图)时才逐级提高阈值兜底, 避免卡死。

    返回: [(y0, y1), ...]
    """
    if threshold is not None:
        return _scan_red_rows(arr, threshold, gap, min_height)
    width = arr.shape[1]
    for thr in (max(12, int(width * 0.024)),
                max(20, int(width * 0.04)),
                max(32, int(width * 0.07)),
                max(60, int(width * 0.12))):
        bands = _scan_red_rows(arr, thr, gap, min_height)
        if len(bands) <= max_bands:
            return bands
    # 仍过多: 用很高阈值兜底并截断, 保证不触发海量 Vision 调用
    return _scan_red_rows(arr, max(80, int(width * 0.16)), gap, min_height)[:max_bands]


def _call_title_vision(img_b64):
    """识别单个红色条的板块名和数量，返回 (板块名str, 数量int|None) | (None, None)"""
    api_key = get_api_key()
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}
    payload = {
        "model": get_vision_model(),
        "messages": [{"role": "user", "content": [
            {"type": "image_url", "image_url": {"url": f"data:image/png;base64,{img_b64}"}},
            {"type": "text", "text": _TITLE_PROMPT},
        ]}],
        "thinking": {"type": "disabled"},
        "temperature": 0.01,
        "max_tokens": 1024,
    }
    for attempt in range(1, 6):
        try:
            resp = requests.post(GLM_CHAT_URL, headers=headers, json=payload, timeout=60)
            result = resp.json()
        except Exception:
            if attempt < 5:
                time.sleep(2 ** (attempt - 1))
                continue
            return None, None
        try:
            msg = result['choices'][0]['message']
            content = msg.get('content')
            if isinstance(content, list):
                content = ''.join(p.get('text', '') for p in content if isinstance(p, dict))
            # content 为空时(推理模型把token耗在思考上)，回退到 reasoning_content
            if not content:
                content = msg.get('reasoning_content', '') or ''
        except (KeyError, IndexError, TypeError):
            # 限速重试（智谱限流码 1302 / HTTP 429）
            if result.get('code') in (1302, 429) or '429' in str(result):
                if attempt < 5:
                    time.sleep(10 * attempt)
                    continue
            return None, None
        # 解析 JSON 对象
        m = re.search(r'\{[^}]*\}', content)
        if m:
            try:
                obj = json.loads(m.group(0))
                name = str(obj.get('板块', '')).strip()
                count = obj.get('数量')
                try:
                    count = int(count)
                except (TypeError, ValueError):
                    count = None
                if name:
                    return name, count
            except Exception:
                pass
        # JSON 解析失败时，从文本中兜底提取（优先取带【】的真实标题，
        # 取最后一个匹配，避开思考过程中复述 prompt 的示例文字）
        matches = re.findall(r'【([^】\n]{1,12})】\s*(\d+)\s*只', content)
        if matches:
            name, cnt = matches[-1]
            return name.strip(), int(cnt)
        return None, None
    return None, None


def extract_sectors_by_red(image_path, layout_details, use_cache=True):
    """
    板块识别主方案（红色定位标题+数量 + OCR按序切分）

    流程:
        1. 红色检测定位所有板块标题的 y 坐标
        2. 裁剪每个红色条小图，Vision 识别板块名 + 数量N（小图准、省token）
        3. OCR 按 y 顺序提取所有股票
        4. 按每个板块声明的数量N顺序切分填充（绕开坐标匹配误差）
        5. 遇到"涨停炸板"标题后的股票全部丢弃

    返回: (sectors: list, error: str|None)
    """
    from PIL import Image
    import numpy as np
    import io

    cache_path = image_path + '.sectors.json'

    # 1. 红色检测板块标题 y 坐标 + 识别板块名和数量（带缓存）
    # 缓存格式: [[y, 板块名, 数量], ...]
    sector_titles = None
    if use_cache and os.path.exists(cache_path):
        try:
            if os.path.getmtime(cache_path) >= os.path.getmtime(image_path):
                with open(cache_path, 'r', encoding='utf-8') as f:
                    cached = json.load(f)
                # 兼容旧缓存(无数量) -> 视为无效，重新识别
                if cached and all(len(t) >= 3 for t in cached):
                    sector_titles = cached
        except Exception:
            sector_titles = None

    im = Image.open(image_path).convert('RGB')
    w, h = im.size

    # 校验元信息
    meta = {'title_total': 0, 'title_fail': 0, 'issues': []}

    if sector_titles is None:
        arr = np.array(im)
        bands = _find_red_bands(arr)
        print(f"    红色检测: {len(bands)} 个板块标题")
        sector_titles = []
        fail_count = 0
        for bi, (y0, y1) in enumerate(bands):
            crop = im.crop((0, max(0, y0 - 5), w, min(h, y1 + 5)))
            buf = io.BytesIO()
            crop.save(buf, format='PNG')
            b64 = base64.b64encode(buf.getvalue()).decode()
            name, count = _call_title_vision(b64)
            # 识别失败再重试一次（偶发输出异常）
            if not name:
                time.sleep(1.0)
                name, count = _call_title_vision(b64)
            if name:
                sector_titles.append([y0, name, count])
            else:
                # 仍失败：保留占位(板块名None)，避免打乱按数量切分的顺序
                sector_titles.append([y0, None, None])
                fail_count += 1
            time.sleep(0.5)  # 节流，规避限速
        # 仅当识别基本完整时才写缓存，避免坏结果被固化
        if fail_count == 0:
            try:
                with open(cache_path, 'w', encoding='utf-8') as f:
                    json.dump(sector_titles, f, ensure_ascii=False)
            except Exception:
                pass
        else:
            print(f"    警告: {fail_count} 个板块标题识别失败，未写缓存")

    if not sector_titles:
        return None, "未检测到板块标题", meta

    # 标题识别失败数统计
    meta['title_total'] = len(sector_titles)
    meta['title_fail'] = sum(1 for t in sector_titles if t[1] is None)
    if meta['title_fail'] > 0:
        meta['issues'].append(
            f"{meta['title_fail']}/{meta['title_total']} 个板块标题识别失败"
            f"（板块名/数量缺失，相关股票归属可能不准）")

    # 2. 提取所有股票（按 y 排序）
    stocks = _extract_stocks_with_y(layout_details)
    if not stocks:
        return None, "未提取到股票", meta

    # 3. 按 y 排序板块标题，找"涨停炸板"作为截止线
    titles = sorted(sector_titles, key=lambda x: x[0])
    # 记录所有识别成功的板块标题名（供04图判定：是否含【市场连板股】）
    meta['titles'] = [t[1] for t in titles if t[1]]
    cutoff_idx = len(titles)
    for i, t in enumerate(titles):
        if t[1] in ('涨停炸板', '炸板'):
            cutoff_idx = i
            break
    valid_titles = titles[:cutoff_idx]  # 炸板及之后的板块丢弃

    # 4. 按数量顺序切分填充股票
    #    - tcount 已知: 直接取 tcount 只
    #    - tcount 缺失(含识别失败占位): 用下一个标题 y 坐标界定该段范围
    results = []
    idx = 0
    n_stocks = len(stocks)
    for ti, (ty, tname, tcount) in enumerate(valid_titles):
        # 页面大标题(如"湖南人涨停复盘")不是板块，不吸收任何股票
        if tname in PAGE_TITLES:
            continue
        # 数量缺失时，用下一个标题的 y 坐标界定本段股票数
        if tcount is None:
            next_y = valid_titles[ti + 1][0] if ti + 1 < len(valid_titles) else None
            cnt = 0
            j = idx
            while j < n_stocks and (next_y is None or stocks[j]['_y'] < next_y):
                cnt += 1
                j += 1
            tcount = cnt
        seg = stocks[idx:idx + tcount]
        # 板块声明数量与实际可取股票数不符（股票被截断）
        if tname is not None and not _is_excluded_sector(tname) and len(seg) < tcount:
            meta['issues'].append(
                f"板块「{tname}」声明{tcount}只，实际仅{len(seg)}只（股票数量不足/被截断）")
        idx += tcount
        # 识别失败(tname=None)或排除板块: 仅推进索引，不写入结果
        if tname is None or _is_excluded_sector(tname):
            continue
        for s in seg:
            results.append({
                '概念板块': tname,
                '代码': s['代码'],
                '名称': s['名称'],
                '末次时间': s['末次时间'],
                '连扳数': s['连扳数'],
                '原因': s['原因'],
            })

    # 剩余未分配股票校验: 只统计落在"有效板块(非排除)区域"内但未被吸收的股票。
    # 排除板块(其他热点/其他个股/涨停炸板等)区域的股票本就该丢弃, 不算未分配。
    # 有效区域 = 第一个非排除板块标题 y ~ 最后一个非排除板块的下一个标题 y
    non_excluded = [t for t in valid_titles
                    if t[1] and not _is_excluded_sector(t[1]) and t[1] not in PAGE_TITLES]
    leftover = 0
    if non_excluded:
        region_start_y = non_excluded[0][0]
        # 有效区域结束 = 最后一个非排除板块之后的下一个标题 y(排除板块或涨停炸板)
        region_end_y = float('inf')
        last_ne_y = non_excluded[-1][0]
        for t in titles:
            if t[0] > last_ne_y:
                region_end_y = t[0]
                break
        leftover = sum(1 for s in stocks[idx:]
                       if region_start_y <= s['_y'] < region_end_y)
    if leftover > 3:
        meta['issues'].append(
            f"有 {leftover} 只股票未分配到板块（数量切分与实际不符）")

    return results, None, meta


# 字段识别正则
_TIME_RE = re.compile(r'^\d{1,2}:\d{2}$')          # 末次时间 HH:MM
_CODE_RE = re.compile(r'^\d{6}$')                  # 6位股票代码
# 连板列: 纯数字 / "N板" / "X/Y天Z板" / "Y天Z板" / 涨跌幅百分数(如"19.52%")
#   注: 板块个股表(Schema B)第4列为涨跌幅, 非连板数; "1"=首板, "19.52%"=未涨停大涨股
#   百分数也纳入匹配, 统一写入连扳数字段(Excel仅有该列, 涨跌幅不单列)
_LIANBAN_RE = re.compile(r'^\d+$|^\d+板$|^\d+天\d+板$|\d+/\d+天\d+板|^\d+(?:\.\d+)?%$')
# 表头/非数据 token (含Schema B表头: 首次时间/涨跌幅, 避免泄漏到上一行原因)
_HEADER_TOKENS = {'代码', '名称', '板块', '末次时间', '时间', '连板数', '连扳数', '原因',
                  '首次时间', '涨跌幅'}


def _extract_stocks_with_y(layout_details):
    """
    提取所有股票并带上 y 坐标，按 y 排序。

    兼容两种 OCR 表格结构（长图底部常出现结构退化）:
      A. 规整行: 一行5列  代码|名称|时间|连板|原因
      B. 退化行: 每个字段被拆成单列单行

    做法: 把所有单元格按行序拉平成 (token, y) 流，以6位代码为边界切分成股票，
          遇到【板块】标题或"消息面"说明则中断当前股票（避免跨板块粘连）。
    """
    # 1. 拉平所有单元格为 (文本, y) token 流
    tokens = []
    for page in layout_details:
        for item in page:
            if item.get('label') != 'table':
                continue
            bbox = item.get('bbox_2d', [0, 0, 0, 0])
            y_base = bbox[1] if len(bbox) >= 2 else 0
            th = (bbox[3] - bbox[1]) if len(bbox) >= 4 else 0
            soup = BeautifulSoup(item.get('content', ''), 'html.parser')
            for table in soup.find_all('table'):
                trs = table.find_all('tr')
                n = len(trs)
                for ri, row in enumerate(trs):
                    # 估算该行 y（表格内按行均分）
                    row_y = y_base + int(th * ri / max(n, 1)) if th else y_base
                    for cell in row.find_all(['th', 'td']):
                        t = cell.get_text(strip=True)
                        if t:
                            tokens.append((t, row_y))

    # 2. 以6位代码为边界切分股票
    stocks = []

    def _flush(cur):
        if not cur:
            return
        name = time = lianban = ''
        reasons = []
        for t in cur['toks']:
            if not time and _TIME_RE.match(t):
                time = t
            elif not name:
                name = t
            elif not lianban and _LIANBAN_RE.match(t):
                lianban = t
            else:
                reasons.append(t)
        stocks.append({
            '代码': cur['code'],
            '名称': name,
            '末次时间': time,
            '连扳数': lianban,
            '原因': ' '.join(reasons),
            '_y': cur['y'],
        })

    cur = None
    for t, y in tokens:
        if t in _HEADER_TOKENS:
            continue
        # 板块标题 / 消息面说明 -> 中断当前股票，不跨板块粘连
        if _SECTOR_HEAD_RE.search(t) or t.startswith('消息面') or t.startswith('消息'):
            _flush(cur)
            cur = None
            continue
        if _CODE_RE.match(t):
            _flush(cur)
            cur = {'code': t, 'toks': [], 'y': y}
        elif cur is not None:
            cur['toks'].append(t)
    _flush(cur)

    stocks.sort(key=lambda s: s['_y'])
    return stocks


# ============== 找 03/04 图片 ==============

def read_order(date_folder, base_dir='dataresource'):
    """读取 _order.txt，返回 {序号: 文件名}"""
    order_file = os.path.join(base_dir, date_folder, '_order.txt')
    order = {}
    if not os.path.exists(order_file):
        return order
    with open(order_file, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            m = re.match(r'^(\d+)\.\s*(\S+\.png)$', line, re.I)
            if m:
                order[int(m.group(1))] = m.group(2)
    return order


def locate_03_04(date_folder, base_dir='dataresource'):
    """
    直接从 _order.txt 的第3/4张图开始判定并提取（不轮询，省时省token）。
    找图与提取合并在一步完成，03/04 天然来自同一日期文件夹。

    03判定: OCR 后开头200字含 "880005" 或 "涨跌家数" -> 提取涨跌家数/成交额
    04判定: extract_sectors_by_red 识别到的板块标题含 "市场连板股" -> 提取板块个股

    返回:
        {
          'date':      日期文件夹,
          'has_03':    bool,  'img_03': path|None,
          'zhangdie':  {上涨家数, 下跌家数, 总成交额},
          'reason_03': str|None,          # 03不可用的原因
          'has_04':    bool,  'img_04': path|None,
          'sectors':   list|None, 'meta': dict,
          'reason_04': str|None,          # 04不可用的原因
        }
    """
    res = {
        'date': date_folder,
        'has_03': False, 'img_03': None,
        'zhangdie': {'上涨家数': None, '下跌家数': None, '总成交额': None},
        'reason_03': None,
        'has_04': False, 'img_04': None,
        'sectors': None, 'meta': {},
        'reason_04': None,
    }

    folder = os.path.join(base_dir, date_folder)
    if not os.path.isdir(folder):
        res['reason_03'] = res['reason_04'] = f"文件夹不存在: {folder}"
        return res

    order = read_order(date_folder, base_dir)

    # ---- 03: 直接取 _order.txt 第3张 ----
    if 3 not in order:
        res['reason_03'] = "无 _order.txt 第3张记录"
    else:
        name3 = order[3]
        path3 = os.path.join(folder, name3)
        if not os.path.exists(path3):
            res['reason_03'] = f"03候选图不存在: {name3}"
        else:
            layout3, err3 = glm_ocr(path3)
            if err3:
                res['reason_03'] = f"03图OCR失败: {err3}"
            else:
                head = layout_to_text(layout3, limit=200)
                if '880005' in head or '涨跌家数' in head:
                    res['has_03'] = True
                    res['img_03'] = path3
                    res['zhangdie'] = extract_zhangdie(layout3)
                    print(f"  03图(order#3命中): {name3}")
                else:
                    res['reason_03'] = f"03候选图({name3})开头未含 880005/涨跌家数"

    # ---- 04: 直接取 _order.txt 第4张 ----
    if 4 not in order:
        res['reason_04'] = "无 _order.txt 第4张记录"
    else:
        name4 = order[4]
        path4 = os.path.join(folder, name4)
        if not os.path.exists(path4):
            res['reason_04'] = f"04候选图不存在: {name4}"
        else:
            layout4, err4 = glm_ocr(path4)
            if err4:
                res['reason_04'] = f"04图OCR失败: {err4}"
            else:
                sectors, errsec, meta = extract_sectors_by_red(path4, layout4)
                res['meta'] = meta or {}
                titles = (meta or {}).get('titles', [])
                has_market = any('市场连板' in (t or '') for t in titles)
                if errsec:
                    res['reason_04'] = f"04图板块识别失败: {errsec}"
                elif not has_market:
                    res['reason_04'] = f"04候选图({name4})未识别到【市场连板股】，判定非04图"
                else:
                    res['has_04'] = True
                    res['img_04'] = path4
                    res['sectors'] = sectors or []
                    print(f"  04图(order#4命中): {name4}")

    return res


# ============== 单日提取 ==============

def export_excel(date_folder, zhangdie, sectors, output_dir='excelDataSource',
                 status='verified', issues=None,
                 has_03=True, has_04=True, reason_03=None, reason_04=None):
    """
    基于提取模板生成Excel（03/04 只要命中其一即生成，缺失的一方在表内写明原因）
    - 03提取 sheet: 上涨家数 | 下跌家数 | 总成交额
    - 04提取 sheet: 概念板块 | 代码 | 名称 | 末次时间 | 连扳数 | 原因
    - 文件名按校验状态加后缀: _verified(通过) / _manualcheck(需人工复核)
    - 缺失的 03/04 在对应 sheet 内写明原因；04提取 sheet 末尾追加校验说明
    """
    os.makedirs(output_dir, exist_ok=True)
    issues = issues or []
    sectors = sectors or []
    zhangdie = zhangdie or {}

    # 以模板为基础（保留表头/样式）
    if os.path.exists(TEMPLATE_PATH):
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
    else:
        wb = openpyxl.Workbook()
        wb.active.title = '03提取'
        wb['03提取'].append(['上涨家数', '下跌家数', '总成交额'])
        ws04 = wb.create_sheet('04提取')
        ws04.append(['概念板块', '代码', '名称', '末次时间', '连扳数', '原因'])

    # === 03提取 ===
    ws03 = wb['03提取']
    if has_03:
        ws03.cell(row=2, column=1, value=zhangdie.get('上涨家数'))
        ws03.cell(row=2, column=2, value=zhangdie.get('下跌家数'))
        ws03.cell(row=2, column=3, value=zhangdie.get('总成交额'))
    else:
        # 未找到03图：留空数据行并写明原因
        ws03.append([])
        ws03.append(['【未找到03图】', reason_03 or '未找到符合条件的03图片'])

    # === 04提取 ===
    ws04 = wb['04提取']
    # 清除模板里第2行起的旧数据（保留表头）
    if ws04.max_row > 1:
        ws04.delete_rows(2, ws04.max_row)
    if has_04:
        for s in sectors:
            ws04.append([
                s.get('概念板块', ''),
                s.get('代码', ''),
                s.get('名称', ''),
                s.get('末次时间', ''),
                s.get('连扳数', ''),
                s.get('原因', ''),
            ])
    else:
        # 未找到04图：写明原因
        ws04.append(['【未找到04图】', reason_04 or '未找到符合条件的04图片'])

    # === 末尾追加校验说明 ===
    ws04.append([])
    if status == 'verified':
        ws04.append(['【校验结果】', '通过 (verified)'])
        ws04.append(['说明', '03/04均识别成功，板块标题与声明数量校验通过。'])
    else:
        ws04.append(['【校验结果】', '需人工复核 (manualcheck)'])
        for i, msg in enumerate(issues, 1):
            ws04.append([f'问题{i}', msg])
        ws04.append(['建议', '请对照原始图片人工核对上述问题（含缺失的03/04图）。'])

    suffix = 'verified' if status == 'verified' else 'manualcheck'
    out_path = os.path.join(output_dir, f'{date_folder}_涨停复盘_{suffix}.xlsx')
    wb.save(out_path)
    return out_path


# ============== 单日提取 ==============

def extract_date(date_folder, base_dir='dataresource', output_dir='excelDataSource'):
    """提取单日数据并生成Excel（找图与提取合并，直接从 _order.txt 第3/4张开始）"""
    date_folder = date_folder.replace('-', '')
    print("=" * 60)
    print(f"提取日期: {date_folder}")
    print("=" * 60)

    # 1. 从 _order.txt 第3/4张直接判定并提取（不轮询）
    print("[1/2] 从 _order.txt 第3/4张直接判定并提取...")
    r = locate_03_04(date_folder, base_dir)

    # 03/04 都不存在才失败
    if not r['has_03'] and not r['has_04']:
        print("✗ 找不到03/04图片")
        print(f"    03: {r['reason_03']}")
        print(f"    04: {r['reason_04']}")
        return False

    zhangdie = r['zhangdie']
    sectors = r['sectors'] or []

    # 汇总校验问题，判定状态
    issues = list(r['meta'].get('issues', [])) if r['has_04'] else []
    if r['has_03']:
        print(f"  涨跌(03): {zhangdie}")
        if zhangdie.get('上涨家数') is None or zhangdie.get('下跌家数') is None:
            issues.append("03图涨跌家数提取不完整（上涨/下跌家数缺失）")
        if zhangdie.get('总成交额') is None:
            issues.append("03图总成交额未提取到")
    else:
        issues.append(f"未找到03图: {r['reason_03']}")

    if r['has_04']:
        from collections import Counter
        dist = Counter(s['概念板块'] for s in sectors)
        print(f"  板块个股(04): {len(sectors)} 条，{len(dist)} 个板块")
        print(f"  板块分布: {dict(dist)}")
    else:
        issues.append(f"未找到04图: {r['reason_04']}")

    # 仅当 03/04 都命中且无其他问题时才判定通过
    status = 'verified' if (r['has_03'] and r['has_04'] and not issues) else 'manualcheck'
    if status == 'verified':
        print("  ✓ 校验通过")
    else:
        print(f"  ⚠ 校验未通过 ({len(issues)} 项问题):")
        for m in issues:
            print(f"      - {m}")

    # 2. 生成Excel（只要命中其一即生成，缺失方在表内写明原因）
    print("[2/2] 生成Excel(按模板)...")
    excel_path = export_excel(date_folder, zhangdie, sectors, output_dir,
                              status=status, issues=issues,
                              has_03=r['has_03'], has_04=r['has_04'],
                              reason_03=r['reason_03'], reason_04=r['reason_04'])
    print(f"✓ 已生成: {excel_path}")
    print("=" * 60)
    return True


def list_dates(base_dir='dataresource', prefix=None):
    """列出 dataresource 下的日期文件夹（可按前缀过滤），升序"""
    if not os.path.isdir(base_dir):
        return []
    dates = [d for d in os.listdir(base_dir)
             if os.path.isdir(os.path.join(base_dir, d)) and re.match(r'^\d{8}$', d)]
    if prefix:
        dates = [d for d in dates if d.startswith(prefix)]
    return sorted(dates)


def extract_batch(dates, base_dir='dataresource', output_dir='excelDataSource',
                  skip_existing=True):
    """批量提取多个日期"""
    ok, skipped, failed = [], [], []
    for i, d in enumerate(dates, 1):
        d = d.replace('-', '')
        # 已存在任一后缀的导出文件则跳过
        existing = [f for f in os.listdir(output_dir)
                    if f.startswith(f'{d}_涨停复盘') and f.endswith('.xlsx')] \
            if os.path.isdir(output_dir) else []
        if skip_existing and existing:
            print(f"[{i}/{len(dates)}] {d} 已存在Excel({existing[0]})，跳过")
            skipped.append(d)
            continue
        print(f"\n[{i}/{len(dates)}] 处理 {d}")
        try:
            if extract_date(d, base_dir, output_dir):
                ok.append(d)
            else:
                failed.append(d)
        except Exception as e:
            print(f"✗ {d} 异常: {e}")
            failed.append(d)

    print("\n" + "=" * 60)
    print(f"批量完成: 成功 {len(ok)} | 跳过 {len(skipped)} | 失败 {len(failed)}")
    if failed:
        print(f"失败日期: {failed}")
    print("=" * 60)
    return ok, skipped, failed


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("用法:")
        print("  py extract_glm.py <日期>            单日,如 20260601")
        print("  py extract_glm.py <日期> <日期> ... 多日")
        print("  py extract_glm.py month 202606      按月批量(自动跳过已导出)")
        print("  py extract_glm.py all               全部日期批量")
        sys.exit(1)

    if sys.argv[1] == 'month' and len(sys.argv) >= 3:
        dates = list_dates(prefix=sys.argv[2])
        print(f"按月 {sys.argv[2]}: 共 {len(dates)} 天")
        extract_batch(dates)
    elif sys.argv[1] == 'all':
        dates = list_dates()
        print(f"全部: 共 {len(dates)} 天")
        extract_batch(dates)
    else:
        for d in sys.argv[1:]:
            extract_date(d)
            print()
