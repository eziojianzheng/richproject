#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
涨停复盘数据提取API
从03、04图片中提取数据并生成Excel
"""

import os
import re
import json
import yaml
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
from flask import Flask, request, jsonify

from utils import ocr_image_text, find_image_by_content

app = Flask(__name__)

# 加载配置
def load_config():
    """加载配置文件"""
    config_path = 'config.yml'
    if os.path.exists(config_path):
        with open(config_path, 'r', encoding='utf-8') as f:
            return yaml.safe_load(f)
    return {}

CONFIG = load_config()

# 排除的板块
EXCLUDED_SECTORS = ['其他', '涨停炸板', '市场连板股']


def get_image_order(date_folder, base_dir='dataresource'):
    """获取图片顺序映射"""
    order_file = os.path.join(base_dir, date_folder, '_order.txt')
    image_order = {}  # {序号: 文件名}
    
    if os.path.exists(order_file):
        with open(order_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and '.png' in line and not line.startswith('#'):
                    parts = line.split()
                    if len(parts) >= 2:
                        try:
                            order_num = int(parts[0].replace('.', ''))
                            img_name = parts[1]
                            image_order[order_num] = img_name
                        except:
                            pass
    return image_order


def find_image_03_04(date_folder, base_dir='dataresource'):
    """
    查找03图片（涨跌家数）和04图片（涨停复盘）
    
    返回:
        {
            'image_03': 路径或None,
            'image_04': 路径或None,
            'date': 日期
        }
    """
    result = {
        'image_03': None,
        'image_04': None,
        'date': date_folder
    }
    
    folder_path = os.path.join(base_dir, date_folder)
    if not os.path.exists(folder_path):
        return result
    
    # 获取图片顺序
    image_order = get_image_order(date_folder, base_dir)
    
    # 查找04图片（涨停复盘）
    # 先检查第04张
    if 4 in image_order:
        img_path = os.path.join(folder_path, image_order[4])
        text = ocr_image_text(img_path)
        if '湖南人涨停复盘' in text:
            result['image_04'] = img_path
    
    # 如果第04张不是，遍历查找
    if not result['image_04']:
        result['image_04'] = find_image_by_content(date_folder, ['湖南人涨停复盘'], base_dir)
    
    # 查找03图片（涨跌家数）
    # 先检查第03张
    if 3 in image_order:
        img_path = os.path.join(folder_path, image_order[3])
        text = ocr_image_text(img_path)
        if '涨跌' in text or '家数' in text:
            result['image_03'] = img_path
    
    # 如果第03张不是，遍历查找
    if not result['image_03']:
        for order_num in [2, 1, 5, 6]:
            if order_num in image_order:
                img_path = os.path.join(folder_path, image_order[order_num])
                text = ocr_image_text(img_path)
                if '涨跌' in text or '家数' in text:
                    result['image_03'] = img_path
                    break
    
    return result


def validate_same_date(image_03_path, image_04_path):
    """
    校验03和04图片是否来自同一日期文件夹
    """
    if not image_03_path or not image_04_path:
        return False, "图片路径不能为空"
    
    # 提取日期文件夹
    dir_03 = os.path.basename(os.path.dirname(image_03_path))
    dir_04 = os.path.basename(os.path.dirname(image_04_path))
    
    if dir_03 != dir_04:
        return False, f"图片不在同一日期文件夹: 03在{dir_03}, 04在{dir_04}"
    
    return True, f"校验通过，日期: {dir_03}"


def extract_zhangdie_data(image_path):
    """
    从03图片提取涨跌数据
    
    返回:
        {
            '上涨家数': int,
            '下跌家数': int,
            '总成交额': str
        }
    """
    result = {
        '上涨家数': None,
        '下跌家数': None,
        '总成交额': None
    }
    
    if not image_path or not os.path.exists(image_path):
        return result
    
    text = ocr_image_text(image_path)
    
    # 提取上涨家数
    up_match = re.search(r'上涨[家数]*[:：\s]*(\d+)', text)
    if up_match:
        result['上涨家数'] = int(up_match.group(1))
    
    # 提取下跌家数
    down_match = re.search(r'下跌[家数]*[:：\s]*(\d+)', text)
    if down_match:
        result['下跌家数'] = int(down_match.group(1))
    
    # 提取总成交额
    amount_match = re.search(r'总成交额[:：\s]*([\d.]+)[亿万]?', text)
    if amount_match:
        result['总成交额'] = amount_match.group(1)
    
    return result


def extract_zhangting_data(image_path):
    """
    从04图片提取涨停数据
    
    返回:
        {
            '涨停': int,
            '炸板': int,
            '跌停': int,
            '首版': int,
            '连板': int
        }
    """
    result = {
        '涨停': None,
        '炸板': None,
        '跌停': None,
        '首版': None,
        '连板': None
    }
    
    if not image_path or not os.path.exists(image_path):
        return result
    
    text = ocr_image_text(image_path)
    
    # 提取首板数 (格式: "首板38家" 或 "首版38")
    sb_match = re.search(r'首[版板][：:\s]*(\d+)', text)
    if sb_match:
        result['首版'] = int(sb_match.group(1))
    
    # 提取连板数 (格式: "连板16k" 或 "连板16家")
    lb_match = re.search(r'连板[：:\s]*(\d+)', text)
    if lb_match:
        result['连板'] = int(lb_match.group(1))
    
    # 提取炸板数 (格式: "炸板31只" 或 在"涨停炸板"附近)
    zb_match = re.search(r'炸板[：:\s]*(\d+)', text)
    if zb_match:
        result['炸板'] = int(zb_match.group(1))
    
    # 提取涨停数 - 需要从标题行解析
    # 标题行格式: "湖南人涨停复盘...首板38家连板16k"
    # 或从"涨停炸板"部分获取
    # 先尝试匹配"涨停XX家"
    zt_match = re.search(r'涨停(\d+)家', text)
    if zt_match:
        result['涨停'] = int(zt_match.group(1))
    else:
        # 如果没有明确标注，涨停数 = 首板 + 连板
        if result['首版'] and result['连板']:
            result['涨停'] = result['首版'] + result['连板']
    
    # 提取跌停数
    dt_match = re.search(r'跌停[：:\s]*(\d+)', text)
    if dt_match:
        result['跌停'] = int(dt_match.group(1))
    
    return result


def extract_sectors_data(image_path):
    """
    从04图片提取板块个股数据
    排除【其他】【涨停炸板】
    
    返回:
        [
            {
                '板块': '半导体',
                '代码': '603956',
                '名称': '威派格',
                '连板数': '',
                '涨停时间': '09:25:03',
                '涨停原因': '液冷泵+六张网建设'
            },
            ...
        ]
    """
    results = []
    
    if not image_path or not os.path.exists(image_path):
        return results
    
    text = ocr_image_text(image_path)
    
    # 按【】分割板块
    # 先找出所有【板块名】
    sector_pattern = r'【([^】]+)】'
    sectors = re.findall(sector_pattern, text)
    
    # 过滤排除的板块
    sectors = [s for s in sectors if s not in EXCLUDED_SECTORS]
    
    # 按板块分割文本
    sector_texts = re.split(sector_pattern, text)
    
    # sector_texts: ['', '半导体', '内容', '其他', '内容', ...]
    current_sector = None
    for i, part in enumerate(sector_texts):
        if i % 2 == 1:  # 奇数位置是板块名
            if part not in EXCLUDED_SECTORS:
                current_sector = part
        elif i % 2 == 0 and i > 0 and current_sector:
            # 偶数位置是板块内容
            # 解析个股
            stocks = parse_stocks_from_text(part, current_sector)
            results.extend(stocks)
    
    return results


def parse_stocks_from_text(text, sector):
    """
    从文本中解析个股信息
    
    格式示例:
    威派格09:25:03液冷泵+六张网建设+智慧水务
    603956威派格09:25:03液冷泵
    """
    results = []
    
    # 匹配模式: 代码(6位数字) + 名称 + 时间 + 原因
    # 或: 名称 + 时间 + 原因
    
    # 模式1: 代码+名称+时间+原因
    pattern1 = r'(\d{6})([^\d\s]{2,6})\s*(\d{2}:\d{2}:\d{2}|\d{2}:\d{2})?\s*(.+?)(?=\d{6}|$)'
    
    # 模式2: 名称+时间+原因
    pattern2 = r'([^\d\s]{2,6})\s*(\d{2}:\d{2}:\d{2}|\d{2}:\d{2})\s*(.+?)(?=\d{6}|$)'
    
    # 尝试匹配
    matches = re.findall(pattern1, text)
    
    for match in matches:
        code, name, time, reason = match
        reason = reason.strip() if reason else ''
        
        # 清理reason中的无关内容
        reason = re.sub(r'\d{6}', '', reason).strip()
        
        results.append({
            '板块': sector,
            '代码': code,
            '名称': name.strip(),
            '连板数': '',
            '涨停时间': time if time else '',
            '涨停原因': reason[:50] if reason else ''  # 限制长度
        })
    
    return results


def call_ai_for_extraction(image_path, extract_type='full'):
    """
    调用 GLM-OCR 模型进行图片数据提取
    
    参数:
        image_path: 图片路径
        extract_type: 提取类型 ('zhangdie' | 'zhangting' | 'sectors' | 'full')
    
    返回:
        提取的数据
    """
    import base64
    import requests
    
    # 读取配置
    ai_config = CONFIG.get('ai', {}).get('zhipu', {})
    api_key = ai_config.get('api_key', '')
    base_url = ai_config.get('base_url', 'https://open.bigmodel.cn/api/paas/v4/')
    
    if not api_key:
        return None, "未配置API Key"
    
    # 读取图片并转base64
    with open(image_path, 'rb') as f:
        img_base64 = base64.b64encode(f.read()).decode()
    
    def call_glm_ocr():
        """调用 GLM-OCR API"""
        url = "https://open.bigmodel.cn/api/paas/v4/layout_parsing"
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }
        
        payload = {
            "model": "glm-ocr",
            "file": f"data:image/png;base64,{img_base64}"
        }
        
        try:
            response = requests.post(url, headers=headers, json=payload, timeout=120)
            result = response.json()
            
            if 'layout_details' in result:
                return result, None
            else:
                return None, result.get('error', {}).get('message', '未知错误')
        except Exception as e:
            return None, str(e)
    
    def parse_html_tables(layout_details):
        """解析 GLM-OCR 返回的 HTML 表格"""
        from bs4 import BeautifulSoup
        
        all_tables = []
        
        for page in layout_details:
            for item in page:
                if item.get('label') == 'table':
                    content = item.get('content', '')
                    
                    # 解析 HTML 表格
                    soup = BeautifulSoup(content, 'html.parser')
                    tables = soup.find_all('table')
                    
                    for table in tables:
                        rows = table.find_all('tr')
                        table_data = []
                        
                        for row in rows:
                            cells = row.find_all(['th', 'td'])
                            row_data = [cell.get_text(strip=True) for cell in cells]
                            if row_data:
                                table_data.append(row_data)
                        
                        if table_data:
                            all_tables.append(table_data)
        
        return all_tables
    
    def extract_sectors_from_tables(tables):
        """从表格中提取板块和个股数据"""
        # 板块顺序（根据OCR识别的图片结构）
        sector_names = [
            '市场连板股', '光通信', '光伏', '机器人', '算力',
            '碳化硅', '生猪', '液冷', '其他热点', '其他个股', '涨停炸板'
        ]
        
        # 排除的板块
        excluded = ['其他', '涨停炸板', '其他热点', '其他个股']
        
        sectors = []
        
        for table_idx, table in enumerate(tables):
            # 根据表格索引确定板块名
            if table_idx < len(sector_names):
                current_sector = sector_names[table_idx]
            else:
                current_sector = f'板块{table_idx + 1}'
            
            # 跳过排除的板块
            if current_sector in excluded:
                continue
            
            for row in table:
                if not row or len(row) < 4:
                    continue
                
                first_cell = str(row[0]).strip() if row else ''
                
                # 跳过表头行
                if first_cell in ['代码', '板块']:
                    continue
                
                # 提取股票数据
                code = first_cell
                
                # 验证代码格式（6位数字）
                if re.match(r'\d{6}', code):
                    name = str(row[1]).strip() if len(row) > 1 else ''
                    time = str(row[2]).strip() if len(row) > 2 else ''
                    board = str(row[3]).strip() if len(row) > 3 else ''
                    reason = str(row[4]).strip() if len(row) > 4 else ''
                    
                    sectors.append({
                        's': current_sector,
                        'c': code,
                        'n': name,
                        't': time,
                        'b': board,
                        'r': reason
                    })
        
        return sectors
    
    def extract_summary_from_text(layout_details):
        """从文本中提取涨跌统计数据"""
        from bs4 import BeautifulSoup
        
        summary = {
            '涨停': None,
            '炸板': None,
            '跌停': None,
            '首版': None,
            '连板': None,
            '上涨家数': None,
            '下跌家数': None,
            '总成交额': None
        }
        
        for page in layout_details:
            for item in page:
                label = item.get('label', '')
                content = item.get('content', '')
                
                # 处理文本内容
                if label == 'text':
                    # 提取涨停数据
                    zt_match = re.search(r'涨停(\d+)家', content)
                    if zt_match:
                        summary['涨停'] = int(zt_match.group(1))
                    
                    zb_match = re.search(r'炸板(\d+)\s*家', content)
                    if zb_match:
                        summary['炸板'] = int(zb_match.group(1))
                    
                    dt_match = re.search(r'跌停(\d+)家', content)
                    if dt_match:
                        summary['跌停'] = int(dt_match.group(1))
                    
                    sb_match = re.search(r'首板(\d+)家', content)
                    if sb_match:
                        summary['首版'] = int(sb_match.group(1))
                    
                    lb_match = re.search(r'连板(\d+)家', content)
                    if lb_match:
                        summary['连板'] = int(lb_match.group(1))
                
                # 处理表格内容 - 03图片的涨跌数据在表格中
                elif label == 'table':
                    soup = BeautifulSoup(content, 'html.parser')
                    table_text = soup.get_text(separator=' ')
                    
                    # 提取上涨家数
                    up_match = re.search(r'上涨家数\s*(\d+)', table_text)
                    if up_match:
                        summary['上涨家数'] = int(up_match.group(1))
                    
                    # 提取下跌家数
                    down_match = re.search(r'下跌家数\s*(\d+)', table_text)
                    if down_match:
                        summary['下跌家数'] = int(down_match.group(1))
                    
                    # 提取总成交额
                    amount_match = re.search(r'总成交额\s*([\d.]+)\s*亿', table_text)
                    if amount_match:
                        summary['总成交额'] = amount_match.group(1)
                    
                    # 表格中也可能有涨停/跌停数据
                    zt_match = re.search(r'其中涨停\s*\|\s*(\d+)', table_text)
                    if zt_match:
                        summary['涨停'] = int(zt_match.group(1))
                    
                    dt_match = re.search(r'其中跌停\s*\|\s*(\d+)', table_text)
                    if dt_match:
                        summary['跌停'] = int(dt_match.group(1))
        
        return summary
    
    # 调用 GLM-OCR
    result, err = call_glm_ocr()
    if err:
        return None, f"OCR识别失败: {err}"
    
    layout_details = result.get('layout_details', [])
    
    # 根据提取类型返回不同数据
    if extract_type == 'sectors':
        tables = parse_html_tables(layout_details)
        stocks = extract_sectors_from_tables(tables)
        
        # 按板块分组
        sectors_map = {}
        for item in stocks:
            sector_name = item['s']
            if sector_name not in sectors_map:
                sectors_map[sector_name] = {'s': sector_name, 'k': []}
            sectors_map[sector_name]['k'].append({
                'c': item['c'],
                'n': item['n'],
                't': item['t'],
                'b': item['b'],
                'r': item['r']
            })
        
        return {'sectors': list(sectors_map.values())}, None
    
    elif extract_type == 'zhangting':
        return extract_summary_from_text(layout_details), None
    
    elif extract_type == 'zhangdie':
        # 03图片的涨跌数据需要从不同位置提取
        return extract_summary_from_text(layout_details), None
    
    else:  # full
        tables = parse_html_tables(layout_details)
        stocks = extract_sectors_from_tables(tables)
        summary = extract_summary_from_text(layout_details)
        
        sectors_map = {}
        for item in stocks:
            sector_name = item['s']
            if sector_name not in sectors_map:
                sectors_map[sector_name] = {'s': sector_name, 'k': []}
            sectors_map[sector_name]['k'].append({
                'c': item['c'],
                'n': item['n'],
                't': item['t'],
                'b': item['b'],
                'r': item['r']
            })
        
        return {
            'zhangting': summary,
            'sectors': list(sectors_map.values())
        }, None


def create_excel_from_data(date, zhangdie_data, zhangting_data, sectors_data, 
                           output_dir='excelDataSource', template_path=None):
    """
    根据提取的数据生成Excel文件
    
    参数:
        date: 日期 (YYYYMMDD)
        zhangdie_data: 涨跌数据
        zhangting_data: 涨停数据
        sectors_data: 板块个股数据
        output_dir: 输出目录
        template_path: 模板路径 (可选)
    
    返回:
        生成的Excel文件路径
    """
    # 创建输出目录
    os.makedirs(output_dir, exist_ok=True)
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    
    # === 涨跌统计 Sheet ===
    ws1 = wb.active
    ws1.title = '涨跌统计'
    
    # 标题
    ws1['A1'] = f'涨停复盘 - {date}'
    ws1['A1'].font = Font(bold=True, size=14)
    
    # 涨跌数据
    ws1['A3'] = '上涨家数'
    ws1['B3'] = zhangdie_data.get('上涨家数', '')
    
    ws1['A4'] = '下跌家数'
    ws1['B4'] = zhangdie_data.get('下跌家数', '')
    
    ws1['A5'] = '总成交额(亿)'
    ws1['B5'] = zhangdie_data.get('总成交额', '')
    
    # 涨停数据
    ws1['A7'] = '涨停'
    ws1['B7'] = zhangting_data.get('涨停', '')
    
    ws1['A8'] = '炸板'
    ws1['B8'] = zhangting_data.get('炸板', '')
    
    ws1['A9'] = '跌停'
    ws1['B9'] = zhangting_data.get('跌停', '')
    
    ws1['A10'] = '首版'
    ws1['B10'] = zhangting_data.get('首版', '')
    
    ws1['A11'] = '连板'
    ws1['B11'] = zhangting_data.get('连板', '')
    
    # 设置列宽
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 15
    
    # === 板块个股 Sheet ===
    ws2 = wb.create_sheet('板块个股')
    
    # 表头
    headers = ['板块', '代码', '名称', '连板数', '涨停时间', '涨停原因']
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
    
    # 数据行
    for row_idx, stock in enumerate(sectors_data, 2):
        ws2.cell(row=row_idx, column=1, value=stock.get('板块', ''))
        ws2.cell(row=row_idx, column=2, value=stock.get('代码', ''))
        ws2.cell(row=row_idx, column=3, value=stock.get('名称', ''))
        ws2.cell(row=row_idx, column=4, value=stock.get('连板数', ''))
        ws2.cell(row=row_idx, column=5, value=stock.get('涨停时间', ''))
        ws2.cell(row=row_idx, column=6, value=stock.get('涨停原因', '') or stock.get('r', ''))
    
    # 设置列宽
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 10
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 40
    
    # 保存文件
    output_path = os.path.join(output_dir, f'{date}_涨停复盘.xlsx')
    wb.save(output_path)
    
    return output_path


# ========== API 路由 ==========

@app.route('/', methods=['GET'])
def index():
    """API首页"""
    return jsonify({
        'name': '涨停复盘数据提取API',
        'version': '1.0.0',
        'description': '从03、04图片中提取数据并生成Excel',
        'endpoints': {
            'POST /api/extract': '提取数据并生成Excel',
            'GET /api/extract/<date>': '提取单个日期的数据',
            'POST /api/preview': '预览将要提取的图片'
        },
        'parameters': {
            'template': '模板路径 (默认: 涨停复盘数据提取模版.xlsx)',
            'image_03': '涨跌家数图片路径 (可选，自动查找)',
            'image_04': '涨停复盘图片路径 (可选，自动查找)',
            'output_dir': '输出目录 (默认: excelDataSource)',
            'use_ai': '是否使用AI提取 (默认: false)'
        }
    })


@app.route('/api/extract', methods=['POST'])
def extract_data():
    """
    提取数据并生成Excel
    
    参数:
        date: 日期 (YYYYMMDD格式) 或
        start_date + end_date: 日期范围
        template: 模板路径 (可选)
        image_03: 指定03图片路径 (可选)
        image_04: 指定04图片路径 (可选)
        output_dir: 输出目录 (默认: excelDataSource)
        use_ai: 是否使用AI提取 (默认: false)
    
    示例:
        POST /api/extract
        {
            "date": "20260626",
            "output_dir": "excelDataSource"
        }
        
        或批量:
        {
            "start_date": "20260620",
            "end_date": "20260626"
        }
    """
    data = request.get_json() or {}
    
    # 获取参数
    single_date = data.get('date')
    start_date = data.get('start_date')
    end_date = data.get('end_date')
    template = data.get('template', '涨停复盘数据提取模版.xlsx')
    image_03 = data.get('image_03')
    image_04 = data.get('image_04')
    output_dir = data.get('output_dir', 'excelDataSource')
    use_ai = data.get('use_ai', False)
    
    # 确定要处理的日期列表
    dates = []
    if single_date:
        dates = [single_date]
    elif start_date and end_date:
        # 生成日期范围
        from datetime import datetime, timedelta
        start = datetime.strptime(start_date, '%Y%m%d') if len(start_date) == 8 else datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y%m%d') if len(end_date) == 8 else datetime.strptime(end_date, '%Y-%m-%d')
        current = start
        while current <= end:
            dates.append(current.strftime('%Y%m%d'))
            current += timedelta(days=1)
    else:
        return jsonify({
            'success': False,
            'error': '请提供 date 参数或 start_date + end_date 参数'
        }), 400
    
    results = []
    
    for date in dates:
        print(f"\n处理日期: {date}")
        
        # 查找图片
        if not image_03 or not image_04:
            images = find_image_03_04(date)
            img_03 = image_03 or images['image_03']
            img_04 = image_04 or images['image_04']
        else:
            img_03 = image_03
            img_04 = image_04
        
        # 校验图片
        if not img_03:
            results.append({
                'date': date,
                'success': False,
                'error': '未找到03图片(涨跌家数)'
            })
            continue
        
        if not img_04:
            results.append({
                'date': date,
                'success': False,
                'error': '未找到04图片(涨停复盘)'
            })
            continue
        
        # 校验同一日期
        valid, msg = validate_same_date(img_03, img_04)
        if not valid:
            results.append({
                'date': date,
                'success': False,
                'error': msg
            })
            continue
        
        print(f"  03图片: {os.path.basename(img_03)}")
        print(f"  04图片: {os.path.basename(img_04)}")
        
        # 提取数据
        if use_ai:
            # 使用AI提取
            zhangdie_data, err1 = call_ai_for_extraction(img_03, 'zhangdie')
            zhangting_result, err2 = call_ai_for_extraction(img_04, 'zhangting')
            sectors_result, err3 = call_ai_for_extraction(img_04, 'sectors')
            
            zhangting_data = zhangting_result or {}
            
            if sectors_result and 'sectors' in sectors_result:
                sectors_data = []
                for sector in sectors_result['sectors']:
                    sector_name = sector.get('s', sector.get('sector', sector.get('板块', '')))
                    stocks = sector.get('k', sector.get('stocks', []))
                    for stock in stocks:
                        sectors_data.append({
                            '板块': sector_name,
                            '代码': stock.get('c', stock.get('code', stock.get('代码', ''))),
                            '名称': stock.get('n', stock.get('name', stock.get('名称', ''))),
                            '连板数': stock.get('b', stock.get('board', stock.get('连板数', ''))),
                            '涨停时间': stock.get('t', stock.get('time', stock.get('涨停时间', ''))),
                            '涨停原因': stock.get('r', stock.get('reason', stock.get('涨停原因', '')))
                        })
            else:
                sectors_data = []
        else:
            # 使用OCR提取
            zhangdie_data = extract_zhangdie_data(img_03)
            zhangting_data = extract_zhangting_data(img_04)
            sectors_data = extract_sectors_data(img_04)
        
        print(f"  涨跌数据: {zhangdie_data}")
        print(f"  涨停数据: {zhangting_data}")
        print(f"  板块数量: {len(sectors_data)}")
        
        # 生成Excel
        try:
            excel_path = create_excel_from_data(
                date, zhangdie_data, zhangting_data, sectors_data,
                output_dir, template
            )
            
            results.append({
                'date': date,
                'success': True,
                'excel_path': excel_path,
                'image_03': img_03,
                'image_04': img_04,
                'zhangdie_data': zhangdie_data,
                'zhangting_data': zhangting_data,
                'sectors_count': len(sectors_data)
            })
            
            print(f"  ✓ 生成Excel: {excel_path}")
            
        except Exception as e:
            results.append({
                'date': date,
                'success': False,
                'error': str(e)
            })
    
    # 统计
    success_count = sum(1 for r in results if r.get('success'))
    
    return jsonify({
        'success': True,
        'total': len(dates),
        'success_count': success_count,
        'output_dir': output_dir,
        'use_ai': use_ai,
        'results': results
    })


@app.route('/api/extract/<date>', methods=['GET'])
def extract_single_date(date):
    """
    提取单个日期的数据
    
    参数:
        date: 日期 (YYYYMMDD)
        output_dir: 输出目录 (query参数)
        use_ai: 是否使用AI (query参数)
    """
    # 标准化日期格式
    date_str = date.replace('-', '')
    
    output_dir = request.args.get('output_dir', 'excelDataSource')
    use_ai = request.args.get('use_ai', 'false').lower() == 'true'
    
    # 调用提取
    result = extract_data.__wrapped__(None)  # 重新包装调用
    
    # 简化：直接处理
    print(f"\n提取日期: {date_str}")
    
    # 查找图片
    images = find_image_03_04(date_str)
    img_03 = images['image_03']
    img_04 = images['image_04']
    
    if not img_03 or not img_04:
        return jsonify({
            'success': False,
            'error': f'未找到图片: 03={img_03 is not None}, 04={img_04 is not None}'
        }), 404
    
    # 校验
    valid, msg = validate_same_date(img_03, img_04)
    if not valid:
        return jsonify({
            'success': False,
            'error': msg
        }), 400
    
    # 提取
    if use_ai:
        zhangdie_data, _ = call_ai_for_extraction(img_03, 'zhangdie')
        zhangting_data, _ = call_ai_for_extraction(img_04, 'zhangting')
        sectors_result, _ = call_ai_for_extraction(img_04, 'sectors')
        
        sectors_data = []
        if sectors_result and 'sectors' in sectors_result:
            for sector in sectors_result['sectors']:
                for stock in sector.get('stocks', []):
                    sectors_data.append({
                        '板块': sector.get('板块', ''),
                        '代码': stock.get('代码', ''),
                        '名称': stock.get('名称', ''),
                        '连板数': stock.get('连板数', ''),
                        '涨停时间': stock.get('涨停时间', ''),
                        '涨停原因': stock.get('涨停原因', '')
                    })
    else:
        zhangdie_data = extract_zhangdie_data(img_03)
        zhangting_data = extract_zhangting_data(img_04)
        sectors_data = extract_sectors_data(img_04)
    
    # 生成Excel
    excel_path = create_excel_from_data(
        date_str, zhangdie_data, zhangting_data, sectors_data, output_dir
    )
    
    return jsonify({
        'success': True,
        'date': date_str,
        'excel_path': excel_path,
        'image_03': img_03,
        'image_04': img_04,
        'zhangdie_data': zhangdie_data,
        'zhangting_data': zhangting_data,
        'sectors_count': len(sectors_data)
    })


@app.route('/api/preview', methods=['POST'])
def preview_extract():
    """
    预览将要提取的图片
    
    参数:
        date: 单个日期 或
        start_date + end_date: 日期范围
    """
    data = request.get_json() or {}
    
    single_date = data.get('date')
    start_date = data.get('start_date')
    end_date = data.get('end_date')
    
    # 确定日期
    dates = []
    if single_date:
        dates = [single_date]
    elif start_date and end_date:
        from datetime import datetime, timedelta
        start = datetime.strptime(start_date.replace('-', ''), '%Y%m%d')
        end = datetime.strptime(end_date.replace('-', ''), '%Y%m%d')
        current = start
        while current <= end:
            dates.append(current.strftime('%Y%m%d'))
            current += timedelta(days=1)
    else:
        return jsonify({
            'success': False,
            'error': '请提供日期参数'
        }), 400
    
    results = []
    for date in dates:
        images = find_image_03_04(date)
        
        valid = False
        msg = ''
        if images['image_03'] and images['image_04']:
            valid, msg = validate_same_date(images['image_03'], images['image_04'])
        
        results.append({
            'date': date,
            'image_03': images['image_03'],
            'image_04': images['image_04'],
            'valid': valid,
            'message': msg
        })
    
    return jsonify({
        'success': True,
        'total': len(dates),
        'preview': results
    })


if __name__ == '__main__':
    print("=" * 60)
    print("涨停复盘数据提取API")
    print("=" * 60)
    print("\n启动服务: http://127.0.0.1:5004")
    print("\nAPI接口:")
    print("  POST /api/extract         - 提取数据并生成Excel")
    print("  GET  /api/extract/<date>  - 提取单个日期")
    print("  POST /api/preview         - 预览图片")
    print("\n示例:")
    print('  curl -X POST http://127.0.0.1:5004/api/export \\')
    print('       -H "Content-Type: application/json" \\')
    print('       -d \'{"date":"20260626"}\'')
    print("\n参数说明:")
    print("  template: 模板路径 (默认: 涨停复盘数据提取模版.xlsx)")
    print("  image_03: 指定03图片路径")
    print("  image_04: 指定04图片路径")
    print("  output_dir: 输出目录 (默认: excelDataSource)")
    print("  use_ai: 是否使用AI提取 (默认: false)")
    print("=" * 60)
    
    app.run(host='0.0.0.0', port=5004, debug=True)
